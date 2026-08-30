from __future__ import annotations

import json
import re
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from capabilities.OCRAcquisitionPipeline.constants import (
    RAW_OCR_DIR,
    REFINED_JSON_DIR,
)
from capabilities.OCRAcquisitionPipeline.session_state import (
    get_selected_raw_ocr_file,
    set_selected_refined_json_file,
)


NA = "NA"
REFINEMENT_VERSION = "1.0"

FIELDS = (
    "Total",
    "Store",
    "Six-Digit SKU",
    "Product",
    "Tax Code",
    "Store Number",
    "Common Name",
    "Sub-Category",
    "Date 1",
    "Price 1",
)

PRICE_PATTERN = re.compile(r"(?<!\d)(\d+[.,]\d{2})(?!\d)")
SKU_PATTERN = re.compile(r"(?<!\d)(\d{6})(?!\d)")
TAX_PATTERN = re.compile(
    r"(?<![A-Za-z])(TLF|TF|LF|FA|FB|NB|F|T)(?![A-Za-z])",
    re.IGNORECASE,
)
DATE_PATTERN = re.compile(
    r"(?<!\d)(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})(?!\d)"
)

STORE_NAMES = (
    ("trader joe", "Trader Joe's"),
    ("publix", "Publix"),
    ("aldi", "Aldi"),
)

CATEGORY_RULES = (
    ("Produce", (
        "apple", "banana", "berry", "berries", "blueberry", "blueberries",
        "strawberry", "strawberries", "zucchini", "broccoli", "tomato",
        "tomatoes", "cilantro", "onion", "onions", "potato", "potatoes",
        "pepper", "peppers", "lettuce", "spinach", "avocado", "orange",
        "lemon", "lime", "pineapple", "watermelon", "cauliflower",
    )),
    ("Bakery", (
        "bread", "muffin", "bagel", "roll", "tortilla", "bun", "bakery",
    )),
    ("Dairy", (
        "milk", "cheese", "yogurt", "yobaby", "half&half", "half and half",
        "cream", "butter", "goat cheese",
    )),
    ("Meat & Seafood", (
        "turkey", "trk", "chicken", "beef", "pork", "salmon", "shrimp",
        "ground", "steak", "kosher",
    )),
    ("Frozen", (
        "frozen", "pizza", "ice cream",
    )),
    ("Breakfast", (
        "waffle", "cereal", "oatmeal", "pancake",
    )),
    ("Beverages", (
        "juice", "water", "soda", "coffee", "tea", "drink",
    )),
    ("Snacks", (
        "cracker", "chips", "cookie", "cookies", "snack", "pretzel",
    )),
    ("Pantry", (
        "pasta", "fusilli", "beans", "rice", "sauce", "flour", "oil",
        "vinegar", "canned", "can ", "soup",
    )),
    ("Prepared Foods", (
        "prepared", "deli", "sandwich", "salad kit",
    )),
    ("Household", (
        "pen", "pens", "paper towel", "toilet paper", "detergent", "cleaner",
        "trash bag", "foil", "battery", "batteries",
    )),
    ("Personal Care", (
        "shampoo", "soap", "toothpaste", "deodorant", "lotion",
    )),
    ("Baby", (
        "baby", "diaper", "wipes", "formula",
    )),
    ("Pet", (
        "dog", "cat", "pet", "kibble",
    )),
)

COMMON_NAME_RULES = (
    (("gf wide pan bread", "gluten free bread"), "Gluten-Free Bread"),
    (("ground trk", "ground turkey"), "Ground Turkey"),
    (("r-strawberries", "strawberries 1 lb"), "Strawberries"),
    (("chickpea fusilli",), "Chickpea Fusilli"),
    (("zucchini",), "Zucchini"),
    (("broccoli crowns",), "Broccoli"),
    (("red onions",), "Red Onions"),
    (("beefsteak tomatoes",), "Tomatoes"),
    (("goat cheese",), "Goat Cheese"),
)


def _load_json(path: Path) -> dict:
    with path.open("r", encoding="utf-8") as file:
        return json.load(file)


def _save_json(path: Path, data: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as file:
        json.dump(data, file, ensure_ascii=False, indent=2)
        file.write("\n")


def _normalize(text: str) -> str:
    return " ".join(
        re.sub(r"[^a-z0-9&]+", " ", text.lower()).split()
    )


def _single_price(text: str) -> str:
    values = [value.replace(",", ".") for value in PRICE_PATTERN.findall(text)]
    return values[0] if len(values) == 1 else NA


def _single_sku(text: str) -> str:
    values = SKU_PATTERN.findall(text)
    return values[0] if len(values) == 1 else NA


def _single_tax(text: str) -> str:
    values = [value.upper() for value in TAX_PATTERN.findall(text)]
    return values[0] if len(values) == 1 else NA


def _clean_date(text: str) -> str:
    match = DATE_PATTERN.search(text)
    if match is None:
        return NA

    month, day, year = match.groups()
    if len(year) == 2:
        year = f"20{year}"

    try:
        value = datetime(
            int(year),
            int(month),
            int(day),
        )
    except ValueError:
        return NA

    return value.strftime("%m/%d/%Y")


def _detect_store(lines: list[dict]) -> tuple[str, float, str]:
    for line in lines:
        text = str(line.get("text", ""))
        normalized = text.lower()

        for needle, store in STORE_NAMES:
            if needle in normalized:
                return (
                    store,
                    0.99,
                    f"Receipt contains store identifier '{text.strip()}'.",
                )

    return NA, 0.0, "No supported store name was detected."


def _store_number_candidates(text: str) -> list[str]:
    patterns = (
        r"\bstore\s*(?:#|no\.?|number)?\s*[:#-]?\s*(\d{2,5})\b",
        r"\bst\s*#\s*(\d{2,5})\b",
        r"\bstore#\s*(\d{2,5})\b",
    )
    values = []

    for pattern in patterns:
        values.extend(
            re.findall(pattern, text, flags=re.IGNORECASE)
        )

    return values


def _detect_store_number(lines: list[dict]) -> tuple[str, float, str]:
    for line in lines:
        text = str(line.get("text", ""))
        values = _store_number_candidates(text)
        if len(values) == 1:
            return (
                values[0],
                0.96,
                f"Store number detected from line '{text.strip()}'.",
            )

    return NA, 0.0, "No reliable store number was detected."


def _detect_receipt_date(lines: list[dict]) -> tuple[str, float, str]:
    for line in lines:
        text = str(line.get("text", ""))
        value = _clean_date(text)
        if value != NA:
            return (
                value,
                0.97,
                f"Receipt date detected from line '{text.strip()}'.",
            )

    return NA, 0.0, "No reliable receipt date was detected."


def _is_address(text: str) -> bool:
    lowered = text.lower()
    hints = (
        " ave", " avenue", " blvd", " boulevard", " road", " rd",
        " street", " st ", " highway", " hwy", " suite", " ste ",
    )
    return any(hint in lowered for hint in hints)


def _classification(line: dict, store: str) -> str:
    text = str(line.get("text", "")).strip()
    normalized = _normalize(text)

    row_classification = line.get("row_classification")
    if isinstance(row_classification, dict) and row_classification.get("is_merchandise"):
        return "purchase_item"

    if store != NA and _normalize(store.replace("'s", "")) in normalized.replace("trader joe s", "trader joe"):
        return "store_name"

    if _store_number_candidates(text):
        return "store_number"

    if _clean_date(text) != NA:
        return "receipt_date"

    if _is_address(" " + text.lower() + " "):
        return "address"

    if "subtotal" in normalized:
        return "subtotal"

    if (
        normalized.startswith("total ")
        or normalized == "total"
        or "amount due" in normalized
        or "balance to pay" in normalized
    ):
        return "receipt_total"

    if "tax" in normalized and _single_price(text) != NA:
        return "tax"

    if any(
        term in normalized
        for term in (
            "visa", "mastercard", "credit card", "debit", "cash tendered",
            "change due", "customer copy", "payment",
        )
    ):
        return "payment"

    if any(
        term in normalized
        for term in (
            "auth", "reference", "transaction", "terminal", "entry mode",
            "entrymode", "approval", "approved", "trace",
        )
    ):
        return "transaction_metadata"

    if (
        re.search(r"\b(?:lb|lbs|oz|kg)\b", text, flags=re.IGNORECASE)
        and (" x " in text.lower() or "/" in text)
    ):
        return "quantity_or_weight_detail"

    if normalized in {"", "-", "*"}:
        return "unknown"

    if any(
        term in normalized
        for term in ("thank you", "welcome", "save money", "receipt")
    ):
        return "header_or_footer"

    return "unknown"


def _purchase_description(line: dict, text: str) -> str:
    support = line.get("component_support")
    if isinstance(support, dict):
        description = support.get("description")
        if isinstance(description, dict):
            value = description.get("value")
            if isinstance(value, str) and value.strip():
                return value.strip()

    working = PRICE_PATTERN.sub(" ", text)
    working = SKU_PATTERN.sub(" ", working)
    working = TAX_PATTERN.sub(" ", working)
    working = re.sub(r"\s+", " ", working).strip()
    return working.strip(" -:;,.|_~'\"()[]{}$") or NA


def _non_purchase_product_guess(text: str) -> str:
    cleaned = re.sub(r"\s+", " ", text).strip()
    return cleaned or NA


def _category(product: str) -> tuple[str, float, str]:
    if product == NA:
        return NA, 0.0, "No product identity is available."

    normalized = " " + _normalize(product) + " "

    for category, terms in CATEGORY_RULES:
        if any(term in normalized for term in terms):
            return (
                category,
                0.78,
                f"Product wording is consistent with the '{category}' category.",
            )

    return NA, 0.0, "Product category is not clear enough to infer."


def _common_name(product: str) -> tuple[str, float, str]:
    if product == NA:
        return NA, 0.0, "No product identity is available."

    normalized = _normalize(product)

    for needles, name in COMMON_NAME_RULES:
        if any(needle in normalized for needle in needles):
            return (
                name,
                0.82,
                "Conservative normalization of receipt product wording.",
            )

    # Single/simple produce-like descriptions are already useful common names.
    category, _, _ = _category(product)
    words = product.split()

    if category == "Produce" and len(words) <= 4:
        cleaned = re.sub(
            r"\b\d+(?:\.\d+)?\s*(?:lb|oz|ct)\b",
            "",
            product,
            flags=re.IGNORECASE,
        )
        cleaned = re.sub(r"\s+", " ", cleaned).strip(" -")
        if cleaned:
            return (
                cleaned.title(),
                0.72,
                "Simple produce description normalized to a common name.",
            )

    return NA, 0.0, "No conservative common-name normalization was available."


def _history_rows() -> list[dict]:
    workbook_path = (
        REFINED_JSON_DIR.parent
        / "database"
        / "shopgraph_purchase_history.xlsx"
    )

    if not workbook_path.exists():
        return []

    try:
        workbook = load_workbook(
            workbook_path,
            data_only=False,
            read_only=True,
        )
    except (OSError, ValueError):
        return []

    if "Purchase History" not in workbook.sheetnames:
        workbook.close()
        return []

    sheet = workbook["Purchase History"]

    headers = {}
    for cell in sheet[1]:
        if cell.value is not None:
            headers[str(cell.value).strip()] = cell.column

    required = {
        "Store",
        "Six-Digit SKU",
        "Product",
        "Common Name",
    }

    detailed_header = (
        "Sub-Category"
        if "Sub-Category" in headers
        else (
            "Category"
            if "Category" in headers
            else None
        )
    )

    if (
        not required.issubset(headers)
        or detailed_header is None
    ):
        workbook.close()
        return []

    rows = []

    for row in range(2, sheet.max_row + 1):
        item = {
            key: str(
                sheet.cell(
                    row=row,
                    column=headers[key],
                ).value
                or NA
            ).strip()
            for key in required
        }
        item["Sub-Category"] = str(
            sheet.cell(
                row=row,
                column=headers[detailed_header],
            ).value
            or NA
        ).strip()
        rows.append(item)

    workbook.close()
    return rows


def _history_support(
    history: list[dict],
    store: str,
    sku: str,
    product: str,
) -> dict | None:
    store_key = _normalize(store)
    product_key = _normalize(product)

    for row in history:
        if (
            store != NA
            and sku != NA
            and _normalize(row["Store"]) == store_key
            and row["Six-Digit SKU"] == sku
        ):
            return row

    if product != NA:
        for row in history:
            if (
                store != NA
                and _normalize(row["Store"]) == store_key
                and _normalize(row["Product"]) == product_key
            ):
                return row

    return None


def _value_and_reason(
    value: str,
    confidence: float,
    reason: str,
) -> tuple[str, float, str]:
    if not value or value == NA:
        return NA, 0.0, reason
    return value, round(confidence, 2), reason


def _refine_line(
    line: dict,
    context: dict,
    history: list[dict],
) -> dict:
    source_text = str(line.get("text", "")).strip()
    classification = _classification(
        line,
        context["store"],
    )
    purchase = classification == "purchase_item"

    confidence_raw = line.get("confidence")
    try:
        source_confidence = max(
            0.0,
            min(1.0, float(confidence_raw) / 100.0),
        )
    except (TypeError, ValueError):
        source_confidence = 0.5

    component_support = line.get("component_support")
    if not isinstance(component_support, dict):
        component_support = {}

    sku = NA
    price = NA
    tax_code = NA

    if purchase:
        sku_support = component_support.get("sku")
        if isinstance(sku_support, dict):
            value = sku_support.get("value")
            if isinstance(value, str) and re.fullmatch(r"\d{6}", value):
                sku = value

        if sku == NA:
            sku = _single_sku(source_text)

        price_support = component_support.get("price")
        if isinstance(price_support, dict):
            value = price_support.get("value")
            if isinstance(value, str) and re.fullmatch(r"\d+[.,]\d{2}", value):
                price = value.replace(",", ".")

        if price == NA:
            price = _single_price(source_text)

        tax_support = component_support.get("tax_code")
        if isinstance(tax_support, dict):
            value = tax_support.get("value")
            if isinstance(value, str) and value.strip():
                tax_code = value.upper()

        if tax_code == NA:
            tax_code = _single_tax(source_text)

        product = _purchase_description(line, source_text)

    else:
        product = _non_purchase_product_guess(source_text)

    store = NA
    store_number = NA
    receipt_date = NA

    if purchase:
        store = context["store"]
        store_number = context["store_number"]
        receipt_date = context["receipt_date"]
    elif classification == "store_name":
        store = context["store"]
    elif classification == "store_number":
        values = _store_number_candidates(source_text)
        store_number = values[0] if len(values) == 1 else NA
    elif classification == "receipt_date":
        receipt_date = _clean_date(source_text)

    total = price if purchase and price != NA else NA

    common_name = NA
    common_confidence = 0.0
    common_reason = "Not applicable to this line."

    category = NA
    category_confidence = 0.0
    category_reason = "Not applicable to this line."

    historical = None
    if purchase:
        historical = _history_support(
            history,
            store,
            sku,
            product,
        )

        if historical is not None:
            if historical.get("Product", NA) != NA and product == NA:
                product = historical["Product"]

            if historical.get("Common Name", NA) != NA:
                common_name = historical["Common Name"]
                common_confidence = 0.94
                common_reason = (
                    "Same-store historical product identity supports this Common Name."
                )

            if historical.get("Sub-Category", NA) != NA:
                category = historical["Sub-Category"]
                category_confidence = 0.94
                category_reason = (
                    "Same-store historical product identity supports this Sub-Category."
                )

        if common_name == NA:
            (
                common_name,
                common_confidence,
                common_reason,
            ) = _common_name(product)

        if category == NA:
            (
                category,
                category_confidence,
                category_reason,
            ) = _category(product)

    fields = {field: NA for field in FIELDS}
    field_confidence = {field: 0.0 for field in FIELDS}
    reasoning = {
        field: "No relevant value was identified for this line."
        for field in FIELDS
    }

    def assign(field: str, value: str, confidence: float, reason: str) -> None:
        value, confidence, reason = _value_and_reason(
            value,
            confidence,
            reason,
        )
        fields[field] = value
        field_confidence[field] = confidence
        reasoning[field] = reason

    assign(
        "Product",
        product,
        max(source_confidence, 0.55),
        (
            "Best editable line-level interpretation."
            if not purchase
            else "Product description isolated from merchandise-row components."
        ),
    )

    if purchase:
        assign(
            "Total",
            total,
            max(source_confidence, 0.75) if total != NA else 0.0,
            (
                "For this single receipt observation, Total equals Price 1."
                if total != NA
                else "Purchase price is not reliable enough to establish Total."
            ),
        )
        assign(
            "Store",
            store,
            context["store_confidence"],
            "Receipt-level store propagated to purchase item.",
        )
        assign(
            "Six-Digit SKU",
            sku,
            max(source_confidence, 0.80) if sku != NA else 0.0,
            (
                "Exact six-digit merchandise SKU detected."
                if sku != NA
                else "No reliable six-digit merchandise SKU."
            ),
        )
        assign(
            "Tax Code",
            tax_code,
            max(source_confidence, 0.70) if tax_code != NA else 0.0,
            (
                "Tax marker is associated with this merchandise row."
                if tax_code != NA
                else "No reliable tax code is associated with this item."
            ),
        )
        assign(
            "Store Number",
            store_number,
            context["store_number_confidence"],
            "Receipt-level store number propagated to purchase item.",
        )
        assign(
            "Common Name",
            common_name,
            common_confidence,
            common_reason,
        )
        assign(
            "Sub-Category",
            category,
            category_confidence,
            category_reason,
        )
        assign(
            "Date 1",
            receipt_date,
            context["receipt_date_confidence"],
            "Receipt transaction date propagated to purchase item.",
        )
        assign(
            "Price 1",
            price,
            max(source_confidence, 0.78) if price != NA else 0.0,
            (
                "Price is directly associated with this merchandise row."
                if price != NA
                else "No reliable merchandise price was identified."
            ),
        )

    else:
        if classification == "store_name":
            assign(
                "Store",
                store,
                context["store_confidence"],
                "This line identifies the receipt store.",
            )
        elif classification == "store_number":
            assign(
                "Store Number",
                store_number,
                0.95 if store_number != NA else 0.0,
                "This line identifies the receipt store number.",
            )
        elif classification == "receipt_date":
            assign(
                "Date 1",
                receipt_date,
                0.96 if receipt_date != NA else 0.0,
                "This line identifies the receipt transaction date.",
            )

        reasoning["Total"] = "Not a purchase item."
        reasoning["Six-Digit SKU"] = "No merchandise SKU is assigned to this non-purchase line."
        reasoning["Tax Code"] = "No merchandise tax code is assigned to this non-purchase line."
        reasoning["Common Name"] = "Common Name is reserved for purchase-item normalization."
        reasoning["Sub-Category"] = "Sub-Category is reserved for purchase-item normalization."
        reasoning["Price 1"] = "No product price is assigned to this non-purchase line."

    line_confidence = max(
        [source_confidence]
        + [
            value
            for value in field_confidence.values()
            if isinstance(value, (int, float))
        ]
    )

    return {
        "line_number": line["line_number"],
        "source_text": source_text,
        "classification": classification,
        "confidence": round(line_confidence, 2),
        "fields": fields,
        "field_confidence": field_confidence,
        "reasoning": reasoning,
        "stage6_metadata": {
            "reconstructed": bool(line.get("reconstructed", False)),
            "recovered_missing_row": bool(
                line.get("recovered_missing_row", False)
            ),
            "numeric_refined": bool(line.get("numeric_refined", False)),
            "normalized_y": line.get("normalized_y", NA),
        },
    }


def get_refined_json_path(raw_ocr_path: str | Path) -> Path:
    raw_path = Path(raw_ocr_path)
    stem = raw_path.stem

    if stem.endswith("_raw_ocr"):
        stem = stem[:-8]

    return REFINED_JSON_DIR / f"{stem}_refined.json"


def _list_raw_ocr_files() -> list[Path]:
    if not RAW_OCR_DIR.exists():
        return []

    return sorted(
        RAW_OCR_DIR.glob("*_raw_ocr.json"),
        key=lambda path: path.name.lower(),
    )


def _choose_raw_ocr_file() -> Path | None:
    files = _list_raw_ocr_files()

    if not files:
        print(
            "\n[INFO] No raw OCR JSON files were found in:"
            f"\n{RAW_OCR_DIR}"
        )
        return None

    print("\nSelect raw OCR file to refine:\n")

    for index, path in enumerate(files, start=1):
        print(f"{index}. {path.name}")

    print("0. Cancel")

    while True:
        choice = input("\nSelect option: ").strip()

        if choice == "0":
            return None

        if choice.isdigit():
            index = int(choice)
            if 1 <= index <= len(files):
                return files[index - 1]

        print("\n[ERROR] Invalid option.")


def refine_json_file(
    raw_ocr_path: str | Path,
) -> Path:
    raw_path = Path(raw_ocr_path).expanduser().resolve()

    if not raw_path.exists():
        raise FileNotFoundError(
            f"Raw OCR JSON does not exist: {raw_path}"
        )

    data = _load_json(raw_path)
    lines = data.get("text")

    if not isinstance(lines, list):
        raise ValueError(
            "Raw OCR JSON must contain a top-level 'text' list."
        )

    validated = []

    for line in lines:
        if not isinstance(line, dict):
            raise ValueError(
                "Every raw OCR line must be a JSON object."
            )

        if not isinstance(line.get("line_number"), int):
            raise ValueError(
                "Every raw OCR line must contain an integer line_number."
            )

        if not isinstance(line.get("text"), str):
            raise ValueError(
                "Every raw OCR line must contain text."
            )

        validated.append(line)

    store, store_confidence, store_reason = _detect_store(validated)
    (
        store_number,
        store_number_confidence,
        store_number_reason,
    ) = _detect_store_number(validated)
    (
        receipt_date,
        receipt_date_confidence,
        receipt_date_reason,
    ) = _detect_receipt_date(validated)

    context = {
        "store": store,
        "store_confidence": store_confidence,
        "store_reason": store_reason,
        "store_number": store_number,
        "store_number_confidence": store_number_confidence,
        "store_number_reason": store_number_reason,
        "receipt_date": receipt_date,
        "receipt_date_confidence": receipt_date_confidence,
        "receipt_date_reason": receipt_date_reason,
    }

    history = _history_rows()

    refined_lines = [
        _refine_line(
            line,
            context,
            history,
        )
        for line in validated
    ]

    destination = get_refined_json_path(raw_path)

    output = {
        "source_raw_ocr": str(raw_path),
        "source_receipt": str(data.get("source_receipt", NA)),
        "refinement_version": REFINEMENT_VERSION,
        "receipt_context": {
            "store": store,
            "store_number": store_number,
            "receipt_date": receipt_date,
        },
        "receipt_context_confidence": {
            "store": round(store_confidence, 2),
            "store_number": round(store_number_confidence, 2),
            "receipt_date": round(receipt_date_confidence, 2),
        },
        "receipt_context_reasoning": {
            "store": store_reason,
            "store_number": store_number_reason,
            "receipt_date": receipt_date_reason,
        },
        "line_count": len(refined_lines),
        "lines": refined_lines,
        "provenance": {
            "rule": (
                "Stage 7 preserves one refined object per Stage-6 line. "
                "Receipt context may propagate Store, Store Number, and Date 1 "
                "to genuine purchase items. Product remains an editable best "
                "guess for every source line. Historical Purchase History may "
                "support identity/Common Name/Sub-Category but never supplies a "
                "current receipt Price 1 or Date 1."
            ),
            "purchase_history_reference_used": bool(history),
        },
    }

    _save_json(destination, output)
    set_selected_refined_json_file(destination)

    return destination.resolve()


def ensure_refined_json() -> Path | None:
    raw_path = get_selected_raw_ocr_file()

    if raw_path is None or not raw_path.exists():
        raw_path = _choose_raw_ocr_file()

    if raw_path is None:
        return None

    return refine_json_file(raw_path)


def run_refine_json() -> None:
    print("\n=== Refine Json File ===\n")

    try:
        output_path = ensure_refined_json()

        if output_path is None:
            return

        data = _load_json(output_path)
        context = data["receipt_context"]

        print(
            "[OK] Refined JSON created:"
            f"\n{output_path}"
        )
        print(
            "\nReceipt context:"
            f"\nStore: {context['store']}"
            f"\nStore Number: {context['store_number']}"
            f"\nDate: {context['receipt_date']}"
            f"\nLines refined: {data['line_count']}"
        )
        print(
            "\nThis refined JSON is now available "
            "to Data Base Builder as its preferred guess source."
        )

    except (
        FileNotFoundError,
        ValueError,
        OSError,
        json.JSONDecodeError,
    ) as error:
        print(f"\n[ERROR] {error}")
