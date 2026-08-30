from __future__ import annotations

from collections import defaultdict
from datetime import date, timedelta
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabel, DataLabelList
from openpyxl.styles import Alignment, Font, PatternFill

from utils.DataBaseBuilder.budget_plans.budget_plan_config import (
    CATEGORY_SCOPE_ALL,
    CATEGORY_SCOPE_SELECTED,
    parse_budget,
    parse_iso_date,
    validate_plan,
)
from utils.DataBaseBuilder.excel.category_manager import (
    load_category_mapping_for_analytics,
)
from utils.DataBaseBuilder.excel.purchase_analytics import (
    CURRENCY_FORMAT,
    DATE_FORMAT,
    _atomic_save,
    _broad_category_observations,
    _flatten_purchase_history,
)
from utils.DataBaseBuilder.excel.purchase_history import (
    PURCHASE_SHEET,
    WORKBOOK_PATH,
    _ensure_purchase_schema,
)


CHART_WIDTH = 24
CHART_HEIGHT = 13
HELPER_START_COLUMN = 28  # AB; kept away from the visible chart/key area
HEADER_FILL = "D9EAF7"
SECTION_FILL = "E2F0D9"


class BudgetPlanAnalyticsError(ValueError):
    """Raised when a Budget Plan worksheet cannot be generated safely."""


def _date_range(start: date, end: date):
    current = start
    while current <= end:
        yield current
        current += timedelta(days=1)


def _ideal_remaining(
    budget: float,
    start_date: date,
    end_date: date,
    current_date: date,
) -> float:
    total_elapsed_days = (end_date - start_date).days

    # A one-day plan has only its deadline point. Treat the ideal budget as fully
    # burned by the end of that day rather than divide by zero.
    if total_elapsed_days == 0:
        return 0.0

    elapsed_days = (current_date - start_date).days
    fraction = min(max(elapsed_days / total_elapsed_days, 0.0), 1.0)
    return budget * (1.0 - fraction)


def _relevant_date(start_date: date, end_date: date, today: date | None = None) -> date:
    today = today or date.today()
    if today < start_date:
        return start_date
    if today > end_date:
        return end_date
    return today


def _resolve_category_scope(plan: dict, current_categories: set[str]) -> set[str]:
    if plan["category_scope"] == CATEGORY_SCOPE_ALL:
        if not current_categories:
            raise BudgetPlanAnalyticsError(
                "Category Manager contains no valid broad Categories."
            )
        return set(current_categories)

    missing = [
        category
        for category in plan["categories"]
        if category not in current_categories
    ]
    if missing:
        raise BudgetPlanAnalyticsError(
            "This Budget Plan references Categories that are no longer valid:\n"
            + "\n".join(f'- "{item}"' for item in missing)
        )

    return set(plan["categories"])


def _filtered_observations(
    observations: list[dict],
    start_date: date,
    end_date: date,
    selected_categories: set[str],
) -> list[dict]:
    return sorted(
        [
            item
            for item in observations
            if start_date <= item["date"] <= end_date
            and item["category"] in selected_categories
        ],
        key=lambda item: (
            item["date"],
            item.get("source_row", 0),
            item.get("history_number", 0),
        ),
    )


def _daily_rows(
    observations: list[dict],
    budget: float,
    start_date: date,
    end_date: date,
) -> list[tuple[date, float, float, float]]:
    spending_by_day = defaultdict(float)
    for item in observations:
        spending_by_day[item["date"]] += float(item["price"])

    cumulative = 0.0
    rows = []
    for day in _date_range(start_date, end_date):
        cumulative += spending_by_day.get(day, 0.0)
        rows.append(
            (
                day,
                _ideal_remaining(budget, start_date, end_date, day),
                budget - cumulative,
                cumulative,
            )
        )
    return rows


def _display_categories(plan: dict, selected_categories: set[str]) -> str:
    if plan["category_scope"] == CATEGORY_SCOPE_ALL:
        return "All Categories"
    return ", ".join(sorted(selected_categories, key=str.casefold))


def _ahead_behind_text(expected_spending: float, actual_spending: float) -> str:
    difference = expected_spending - actual_spending
    if abs(difference) < 0.005:
        return "On pace"
    if difference > 0:
        return f"Ahead by ${difference:,.2f}"
    return f"Behind by ${abs(difference):,.2f}"


def _write_budget_plan_sheet(
    workbook,
    plan: dict,
    rows: list[tuple[date, float, float, float]],
    selected_categories: set[str],
    diagnostics: dict,
    purchase_observation_count: int,
    qualifying_observations: list[dict],
) -> dict:
    sheet_name = plan["worksheet_name"]
    if sheet_name in workbook.sheetnames:
        workbook.remove(workbook[sheet_name])

    sheet = workbook.create_sheet(sheet_name)
    sheet.sheet_view.showGridLines = False

    start_date = parse_iso_date(plan["start_date"])
    end_date = parse_iso_date(plan["end_date"])
    budget = float(parse_budget(plan["budget"]))
    relevant = _relevant_date(start_date, end_date)

    row_by_date = {row[0]: row for row in rows}
    relevant_row = row_by_date[relevant]
    ideal_remaining = relevant_row[1]
    actual_remaining = relevant_row[2]
    actual_spending = relevant_row[3]
    expected_spending = budget - ideal_remaining

    sheet["A1"] = plan["name"]
    sheet["A1"].font = Font(size=18, bold=True)
    sheet["A1"].fill = PatternFill("solid", fgColor=HEADER_FILL)
    sheet.merge_cells("A1:H1")

    metrics = [
        ("Budget Plan", plan["name"]),
        ("Start Date", start_date),
        ("End Date", end_date),
        (
            "Category Scope",
            "All Categories"
            if plan["category_scope"] == CATEGORY_SCOPE_ALL
            else "Selected Categories",
        ),
        ("Categories", _display_categories(plan, selected_categories)),
        ("Budget", budget),
        ("Total Spent", actual_spending),
        ("Remaining Budget", budget - actual_spending),
        ("Relevant Date", relevant),
        ("Ideal Spending Through Relevant Date", expected_spending),
        ("Actual Spending Through Relevant Date", actual_spending),
        ("Ahead / Behind Budget", _ahead_behind_text(expected_spending, actual_spending)),
    ]

    for index, (label, value) in enumerate(metrics, start=3):
        sheet.cell(index, 1, label)
        sheet.cell(index, 1).font = Font(bold=True)
        sheet.cell(index, 1).fill = PatternFill("solid", fgColor=SECTION_FILL)
        sheet.cell(index, 2, value)

    for row_index in (4, 5, 11):
        sheet.cell(row_index, 2).number_format = DATE_FORMAT
    for row_index in (8, 9, 12, 13):
        sheet.cell(row_index, 2).number_format = CURRENCY_FORMAT

    sheet["A16"] = "Budget vs Category Spending"
    sheet["A16"].font = Font(size=14, bold=True)

    # Keep the existing daily calculations available as worksheet helper data.
    helper_col = HELPER_START_COLUMN
    helper_headers = (
        "Date",
        "Ideal Remaining Budget",
        "Actual Remaining Budget",
        "Cumulative Spending",
    )
    for offset, header in enumerate(helper_headers):
        cell = sheet.cell(1, helper_col + offset, header)
        cell.font = Font(bold=True)

    for row_number, data in enumerate(rows, start=2):
        for offset, value in enumerate(data):
            cell = sheet.cell(row_number, helper_col + offset, value)
            if offset == 0:
                cell.number_format = DATE_FORMAT
            else:
                cell.number_format = CURRENCY_FORMAT

    # Spending for the comparison bar uses the same relevant-date cutoff as the
    # worksheet summary, so the stacked Spent bar agrees with Total Spent.
    spending_by_category = defaultdict(float)
    for item in qualifying_observations:
        if item["date"] <= relevant:
            spending_by_category[item["category"]] += float(item["price"])

    ordered_categories = sorted(selected_categories, key=str.casefold)

    # Two-bar chart helper table:
    #
    #                Budget   1   2   3 ...
    #   Budget       $B       0   0   0
    #   Spent         0      $1  $2  $3
    #
    # Because every series shares one stacked group, Excel renders exactly two
    # columns. The first is one solid Budget bar; the second is one stacked bar
    # whose colored segments are the individual Categories.
    comparison_helper_col = helper_col + 5  # Q
    comparison_header_row = 1
    comparison_budget_row = 2
    comparison_spent_row = 3

    sheet.cell(comparison_header_row, comparison_helper_col, "Comparison")
    sheet.cell(comparison_header_row, comparison_helper_col + 1, "Budget")
    sheet.cell(comparison_budget_row, comparison_helper_col, "Budget")
    sheet.cell(comparison_spent_row, comparison_helper_col, "Spent")
    sheet.cell(comparison_budget_row, comparison_helper_col + 1, budget)
    sheet.cell(comparison_spent_row, comparison_helper_col + 1, 0.0)
    sheet.cell(comparison_budget_row, comparison_helper_col + 1).number_format = CURRENCY_FORMAT
    sheet.cell(comparison_spent_row, comparison_helper_col + 1).number_format = CURRENCY_FORMAT

    for code, category in enumerate(ordered_categories, start=1):
        column = comparison_helper_col + 1 + code
        sheet.cell(comparison_header_row, column, str(code))
        sheet.cell(comparison_budget_row, column, 0.0)
        sheet.cell(
            comparison_spent_row,
            column,
            spending_by_category.get(category, 0.0),
        )
        sheet.cell(comparison_budget_row, column).number_format = CURRENCY_FORMAT
        sheet.cell(comparison_spent_row, column).number_format = CURRENCY_FORMAT

    for column in range(
        comparison_helper_col,
        comparison_helper_col + 2 + len(ordered_categories),
    ):
        sheet.cell(comparison_header_row, column).font = Font(bold=True)

    chart = BarChart()
    chart.type = "col"
    chart.grouping = "stacked"
    chart.overlap = 100
    chart.style = 12
    chart.title = "Budget vs Category Spending"
    chart.y_axis.title = "Amount ($)"
    chart.x_axis.title = ""
    chart.width = 18
    chart.height = CHART_HEIGHT
    chart.gapWidth = 70

    chart_data = Reference(
        sheet,
        min_col=comparison_helper_col + 1,
        max_col=comparison_helper_col + 1 + len(ordered_categories),
        min_row=comparison_header_row,
        max_row=comparison_spent_row,
    )
    chart_categories = Reference(
        sheet,
        min_col=comparison_helper_col,
        min_row=comparison_budget_row,
        max_row=comparison_spent_row,
    )
    chart.add_data(chart_data, titles_from_data=True)
    chart.set_categories(chart_categories)

    # Fixed palette keeps the chart and its worksheet key synchronized. The
    # first color is reserved for the single Budget bar. Each Category receives
    # a stable numeric code and a distinct segment color in this worksheet.
    budget_color = "5B9BD5"
    category_colors = (
        "ED7D31",
        "70AD47",
        "A5A5A5",
        "FFC000",
        "4472C4",
        "255E91",
        "9E480E",
        "636363",
        "997300",
        "264478",
        "43682B",
        "8064A2",
        "4BACC6",
        "F79646",
        "92A9CF",
        "95B3D7",
        "C0504D",
        "9BBB59",
        "7F7F7F",
        "806000",
    )

    if chart.series:
        budget_series = chart.series[0]
        budget_series.graphicalProperties.solidFill = budget_color
        budget_series.graphicalProperties.line.solidFill = budget_color

    for index, series in enumerate(chart.series[1:], start=1):
        color = category_colors[(index - 1) % len(category_colors)]
        series.graphicalProperties.solidFill = color
        series.graphicalProperties.line.solidFill = color

        # Label only the Spent point (point index 1). The label is the Category
        # number, which corresponds to the color-coded key to the right.
        series.dLbls = DataLabelList(
            dLbl=[
                DataLabel(
                    idx=1,
                    showSerName=True,
                    dLblPos="ctr",
                )
            ]
        )

    # The built-in legend would only repeat the numeric series codes. A custom
    # worksheet key is clearer because it can show number, color, Category name,
    # and current spend together.
    chart.legend = None
    sheet.add_chart(chart, "A18")

    key_start_row = 18
    key_number_col = 10  # J
    key_name_col = 11    # K
    key_spend_col = 12   # L

    sheet.cell(key_start_row, key_number_col, "Category Key")
    sheet.cell(key_start_row, key_number_col).font = Font(size=12, bold=True)
    sheet.merge_cells(
        start_row=key_start_row,
        start_column=key_number_col,
        end_row=key_start_row,
        end_column=key_spend_col,
    )

    sheet.cell(key_start_row + 1, key_number_col, "#")
    sheet.cell(key_start_row + 1, key_name_col, "Category")
    sheet.cell(key_start_row + 1, key_spend_col, "Spent")
    for column in range(key_number_col, key_spend_col + 1):
        sheet.cell(key_start_row + 1, column).font = Font(bold=True)

    for code, category in enumerate(ordered_categories, start=1):
        row_number = key_start_row + 1 + code
        color = category_colors[(code - 1) % len(category_colors)]
        number_cell = sheet.cell(row_number, key_number_col, code)
        number_cell.fill = PatternFill("solid", fgColor=color)
        number_cell.font = Font(bold=True, color="FFFFFF")
        number_cell.alignment = Alignment(horizontal="center")

        sheet.cell(row_number, key_name_col, category)
        spend_cell = sheet.cell(
            row_number,
            key_spend_col,
            spending_by_category.get(category, 0.0),
        )
        spend_cell.number_format = CURRENCY_FORMAT

    # Keep the key readable and separate from the chart. Helper data starts far
    # to the right in column AB, so the visible key can safely use J:L.
    sheet.column_dimensions["A"].width = 38
    sheet.column_dimensions["B"].width = 48
    sheet.column_dimensions["J"].width = 6
    sheet.column_dimensions["K"].width = 28
    sheet.column_dimensions["L"].width = 14

    for column in range(
        helper_col,
        comparison_helper_col + 2 + len(ordered_categories),
    ):
        if column not in (key_number_col, key_name_col, key_spend_col):
            sheet.column_dimensions[
                __import__("openpyxl").utils.get_column_letter(column)
            ].width = 16

    note_row = max(47, key_start_row + 3 + len(ordered_categories))
    sheet.cell(note_row, 1, (
        "Interpretation: the chart has exactly two bars. Budget is the full "
        "configured Budget Plan amount. Spent is one stacked bar; each colored "
        "segment is the amount spent in one Category through the Relevant Date. "
        "The number inside a segment matches the Category Key on the right."
    ))
    sheet.cell(note_row, 1).alignment = Alignment(wrap_text=True, vertical="top")
    sheet.merge_cells(
        start_row=note_row,
        start_column=1,
        end_row=note_row + 2,
        end_column=8,
    )

    return {
        "worksheet_name": sheet_name,
        "budget": budget,
        "total_spent": actual_spending,
        "remaining_budget": actual_remaining,
        "relevant_date": relevant,
        "expected_spending": expected_spending,
        "actual_spending": actual_spending,
        "ahead_behind": _ahead_behind_text(expected_spending, actual_spending),
        "purchase_observations": purchase_observation_count,
        "skipped_pairs": diagnostics.get("skipped_pairs", 0),
        "resolved_categories": sorted(selected_categories, key=str.casefold),
    }


def generate_budget_plan(
    plan: dict,
    workbook_path: Path = WORKBOOK_PATH,
) -> dict:
    plan = validate_plan(plan)
    workbook_path = Path(workbook_path).expanduser().resolve()

    if not workbook_path.exists():
        raise FileNotFoundError(
            "Purchase History workbook does not exist:\n"
            f"{workbook_path}"
        )

    workbook = load_workbook(workbook_path, data_only=False)
    try:
        if PURCHASE_SHEET not in workbook.sheetnames:
            raise BudgetPlanAnalyticsError(
                f'Workbook does not contain "{PURCHASE_SHEET}".'
            )

        purchase_sheet = workbook[PURCHASE_SHEET]
        _ensure_purchase_schema(purchase_sheet)
        observations, diagnostics = _flatten_purchase_history(purchase_sheet)

        mapping = load_category_mapping_for_analytics(workbook)
        broad_observations = _broad_category_observations(observations, mapping)
        current_categories = set(mapping.values())
        selected_categories = _resolve_category_scope(plan, current_categories)

        start_date = parse_iso_date(plan["start_date"])
        end_date = parse_iso_date(plan["end_date"])
        budget = float(parse_budget(plan["budget"]))
        qualifying = _filtered_observations(
            broad_observations,
            start_date,
            end_date,
            selected_categories,
        )
        rows = _daily_rows(qualifying, budget, start_date, end_date)
        summary = _write_budget_plan_sheet(
            workbook,
            plan,
            rows,
            selected_categories,
            diagnostics,
            len(qualifying),
            qualifying,
        )
        _atomic_save(workbook, workbook_path)
        return {
            "success": True,
            "workbook_path": workbook_path,
            **summary,
        }
    finally:
        workbook.close()


def delete_budget_plan_worksheet(
    worksheet_name: str,
    workbook_path: Path = WORKBOOK_PATH,
) -> bool:
    workbook_path = Path(workbook_path).expanduser().resolve()
    if not workbook_path.exists():
        raise FileNotFoundError(
            "Purchase History workbook does not exist:\n"
            f"{workbook_path}"
        )

    workbook = load_workbook(workbook_path, data_only=False)
    try:
        if worksheet_name not in workbook.sheetnames:
            return False
        workbook.remove(workbook[worksheet_name])
        _atomic_save(workbook, workbook_path)
        return True
    finally:
        workbook.close()


def get_current_categories(
    workbook_path: Path = WORKBOOK_PATH,
) -> list[str]:
    workbook_path = Path(workbook_path).expanduser().resolve()
    if not workbook_path.exists():
        raise FileNotFoundError(
            "Purchase History workbook does not exist:\n"
            f"{workbook_path}"
        )

    workbook = load_workbook(workbook_path, data_only=False)
    try:
        mapping = load_category_mapping_for_analytics(workbook)
        return sorted(set(mapping.values()), key=str.casefold)
    finally:
        workbook.close()


def current_worksheet_names(
    workbook_path: Path = WORKBOOK_PATH,
) -> set[str]:
    workbook_path = Path(workbook_path).expanduser().resolve()
    if not workbook_path.exists():
        raise FileNotFoundError(
            "Purchase History workbook does not exist:\n"
            f"{workbook_path}"
        )
    workbook = load_workbook(workbook_path, read_only=True, data_only=False)
    try:
        return set(workbook.sheetnames)
    finally:
        workbook.close()
