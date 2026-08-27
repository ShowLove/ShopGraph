from __future__ import annotations

import os
import tempfile
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from utils.constants import DATA_DIR
from utils.DataBaseBuilder.purchase_record import NA, PurchaseRecord
from utils.DataBaseBuilder.receipt_session import ReceiptSession


DATABASE_DIR = DATA_DIR / "database"
WORKBOOK_PATH = DATABASE_DIR / "shopgraph_purchase_history.xlsx"

PURCHASE_SHEET = "Purchase History"
IMPORT_SHEET = "Imported Receipts"

# ---------------------------------------------------------------------------
# PURCHASE HISTORY SCHEMA
# ---------------------------------------------------------------------------

# New schema:
#
# A  Total
# B  <blank spacer>
# C  Store
# D  Six-Digit SKU
# E  Product
# F  Tax Code
# G  Store Number
# H  Common Name
# I  Category
# J  Date 1
# K  Price 1
# L  Date 2
# M  Price 2
# ...
#
# Common Name and Category are intentionally created as NA. They are reserved
# for future ShopGraph normalization/categorization capabilities.

TOTAL_COLUMN = 1
SPACER_COLUMN = 2
STORE_COLUMN = 3
SKU_COLUMN = 4
PRODUCT_COLUMN = 5
TAX_CODE_COLUMN = 6
STORE_NUMBER_COLUMN = 7
COMMON_NAME_COLUMN = 8
CATEGORY_COLUMN = 9

HISTORY_START_COLUMN = 10

FIXED_HEADERS = [
    "Total",
    "",
    "Store",
    "Six-Digit SKU",
    "Product",
    "Tax Code",
    "Store Number",
    "Common Name",
    "Category",
]

OLD_FIXED_HEADERS = [
    "Six-Digit SKU",
    "Product",
    "Tax Code",
    "Store Number",
]

IMPORT_HEADERS = [
    "Source OCR JSON",
    "Store Number",
    "Receipt Date",
    "Imported Timestamp",
]


# ---------------------------------------------------------------------------
# WORKBOOK CREATION / MIGRATION
# ---------------------------------------------------------------------------

def _create_workbook():
    workbook = Workbook()
    purchase_sheet = workbook.active
    purchase_sheet.title = PURCHASE_SHEET

    for column, header in enumerate(
        FIXED_HEADERS,
        start=1,
    ):
        purchase_sheet.cell(
            row=1,
            column=column,
            value=header,
        )

    _ensure_history_pair(
        purchase_sheet,
        1,
    )

    import_sheet = workbook.create_sheet(
        IMPORT_SHEET
    )

    for column, header in enumerate(
        IMPORT_HEADERS,
        start=1,
    ):
        import_sheet.cell(
            row=1,
            column=column,
            value=header,
        )

    import_sheet.sheet_state = "hidden"

    _format_workbook(
        workbook
    )

    return workbook


def _header_values(
    sheet,
    count: int,
) -> list[str]:
    return [
        str(
            sheet.cell(
                row=1,
                column=column,
            ).value
            or ""
        ).strip()
        for column in range(
            1,
            count + 1,
        )
    ]


def _uses_old_schema(
    sheet,
) -> bool:
    return (
        _header_values(
            sheet,
            len(OLD_FIXED_HEADERS),
        )
        == OLD_FIXED_HEADERS
    )


def _uses_new_schema(
    sheet,
) -> bool:
    return (
        _header_values(
            sheet,
            len(FIXED_HEADERS),
        )
        == FIXED_HEADERS
    )


def _migrate_old_purchase_sheet(
    sheet,
) -> None:
    """
    Convert the original ShopGraph Purchase History schema:

        SKU | Product | Tax Code | Store Number | Date 1 | Price 1 | ...

    into:

        Total | <blank> | Store | SKU | Product | Tax Code |
        Store Number | Common Name | Category | Date 1 | Price 1 | ...

    Existing Date N / Price N history is preserved exactly.
    """
    # Move the original A:D identity columns to D:G.
    sheet.insert_cols(
        1,
        amount=3,
    )

    # After that insertion, the original Date 1 begins at H.
    # Insert Common Name and Category there so history begins at J.
    sheet.insert_cols(
        8,
        amount=2,
    )

    for column, header in enumerate(
        FIXED_HEADERS,
        start=1,
    ):
        sheet.cell(
            row=1,
            column=column,
            value=header,
        )

    for row in range(
        2,
        sheet.max_row + 1,
    ):
        # Existing data does not contain an authoritative store name.
        # Do not guess. A later receipt from the same history can safely
        # populate this field when the selected receipt type is known.
        if sheet.cell(
            row=row,
            column=STORE_COLUMN,
        ).value in (
            None,
            "",
        ):
            sheet.cell(
                row=row,
                column=STORE_COLUMN,
                value=NA,
            )

        existing_common_name = str(
            sheet.cell(
                row=row,
                column=COMMON_NAME_COLUMN,
            ).value
            or NA
        )

        if (
            existing_common_name == NA
            and record.common_name != NA
        ):
            sheet.cell(
                row=row,
                column=COMMON_NAME_COLUMN,
                value=record.common_name,
            )

        existing_category = str(
            sheet.cell(
                row=row,
                column=CATEGORY_COLUMN,
            ).value
            or NA
        )

        if (
            existing_category == NA
            and record.category != NA
        ):
            sheet.cell(
                row=row,
                column=CATEGORY_COLUMN,
                value=record.category,
            )

        _ensure_total_formula(
            sheet,
            row,
        )


def _ensure_purchase_schema(
    sheet,
) -> None:
    if _uses_new_schema(
        sheet
    ):
        for row in range(
            2,
            sheet.max_row + 1,
        ):
            _ensure_total_formula(
                sheet,
                row,
            )
        return

    if _uses_old_schema(
        sheet
    ):
        _migrate_old_purchase_sheet(
            sheet
        )
        return

    raise ValueError(
        "Purchase History has an unrecognized column layout. "
        "Expected either the original ShopGraph schema or the "
        "current ShopGraph schema."
    )


def _load_or_create_workbook():
    if WORKBOOK_PATH.exists():
        workbook = load_workbook(
            WORKBOOK_PATH
        )

        if (
            PURCHASE_SHEET
            not in workbook.sheetnames
        ):
            raise ValueError(
                "Workbook is missing required sheet: "
                f"{PURCHASE_SHEET}"
            )

        purchase_sheet = workbook[
            PURCHASE_SHEET
        ]

        _ensure_purchase_schema(
            purchase_sheet
        )

        if (
            IMPORT_SHEET
            not in workbook.sheetnames
        ):
            import_sheet = workbook.create_sheet(
                IMPORT_SHEET
            )

            for column, header in enumerate(
                IMPORT_HEADERS,
                start=1,
            ):
                import_sheet.cell(
                    row=1,
                    column=column,
                    value=header,
                )

            import_sheet.sheet_state = "hidden"

        return workbook

    return _create_workbook()


# ---------------------------------------------------------------------------
# DATE / PRICE HISTORY
# ---------------------------------------------------------------------------

def _ensure_history_pair(
    sheet,
    pair_number: int,
) -> tuple[int, int]:
    date_column = (
        HISTORY_START_COLUMN
        + (pair_number - 1) * 2
    )

    price_column = (
        date_column + 1
    )

    date_header = sheet.cell(
        row=1,
        column=date_column,
    )

    price_header = sheet.cell(
        row=1,
        column=price_column,
    )

    if date_header.value in (
        None,
        "",
    ):
        date_header.value = (
            f"Date {pair_number}"
        )

    if price_header.value in (
        None,
        "",
    ):
        price_header.value = (
            f"Price {pair_number}"
        )

    return (
        date_column,
        price_column,
    )


def _next_history_pair(
    sheet,
    row: int,
) -> tuple[int, int]:
    pair_number = 1

    while True:
        (
            date_column,
            price_column,
        ) = _ensure_history_pair(
            sheet,
            pair_number,
        )

        date_value = sheet.cell(
            row=row,
            column=date_column,
        ).value

        price_value = sheet.cell(
            row=row,
            column=price_column,
        ).value

        if (
            date_value in (
                None,
                "",
            )
            and price_value in (
                None,
                "",
            )
        ):
            return (
                date_column,
                price_column,
            )

        pair_number += 1


# ---------------------------------------------------------------------------
# TOTAL FORMULA
# ---------------------------------------------------------------------------

def _price_columns(
    sheet,
) -> list[int]:
    """
    Return the worksheet columns that are explicitly labeled Price N.

    In the current Purchase History layout these are:

        K, M, O, Q, S, ...

    Reading the headers keeps the Total formula correct as additional
    Date N / Price N history pairs are added.
    """
    columns = []

    for column in range(
        HISTORY_START_COLUMN + 1,
        sheet.max_column + 1,
        2,
    ):
        header = str(
            sheet.cell(
                row=1,
                column=column,
            ).value
            or ""
        ).strip()

        if header.startswith(
            "Price "
        ):
            columns.append(
                column
            )

    return columns


def _total_formula(
    sheet,
    row: int,
) -> str:
    """
    Sum Price 1, Price 2, Price 3, etc. for one product-history row.

    Example:

        =SUM(K2,M2,O2,Q2,S2)

    Blank future price cells are ignored naturally by Excel.
    """
    price_references = [
        f"{get_column_letter(column)}{row}"
        for column in _price_columns(
            sheet
        )
    ]

    if not price_references:
        return "=0"

    return (
        "=SUM("
        + ",".join(
            price_references
        )
        + ")"
    )


def _ensure_total_formula(
    sheet,
    row: int,
) -> None:
    total_cell = sheet.cell(
        row=row,
        column=TOTAL_COLUMN,
    )

    expected_formula = _total_formula(
        sheet,
        row,
    )

    # If the correct equation is already present, leave it unchanged.
    # If an older/incorrect equation is present, replace it.
    if (
        total_cell.value
        == expected_formula
    ):
        total_cell.number_format = "0.00"
        return

    total_cell.value = (
        expected_formula
    )

    total_cell.number_format = "0.00"


# ---------------------------------------------------------------------------
# FORMATTING
# ---------------------------------------------------------------------------

def _format_workbook(
    workbook,
) -> None:
    purchase_sheet = workbook[
        PURCHASE_SHEET
    ]

    purchase_sheet.freeze_panes = "A2"

    purchase_sheet.auto_filter.ref = (
        purchase_sheet.dimensions
    )

    for cell in purchase_sheet[1]:
        cell.font = Font(
            bold=True
        )

    fixed_widths = {
        TOTAL_COLUMN: 14,
        SPACER_COLUMN: 3,
        STORE_COLUMN: 18,
        SKU_COLUMN: 18,
        PRODUCT_COLUMN: 34,
        TAX_CODE_COLUMN: 14,
        STORE_NUMBER_COLUMN: 16,
        COMMON_NAME_COLUMN: 28,
        CATEGORY_COLUMN: 22,
    }

    for column, width in fixed_widths.items():
        purchase_sheet.column_dimensions[
            get_column_letter(
                column
            )
        ].width = width

    for column in range(
        HISTORY_START_COLUMN,
        purchase_sheet.max_column + 1,
    ):
        purchase_sheet.column_dimensions[
            get_column_letter(
                column
            )
        ].width = 14

    import_sheet = workbook[
        IMPORT_SHEET
    ]

    import_sheet.freeze_panes = "A2"

    for cell in import_sheet[1]:
        cell.font = Font(
            bold=True
        )

    import_widths = [
        32,
        16,
        16,
        22,
    ]

    for column, width in enumerate(
        import_widths,
        start=1,
    ):
        import_sheet.column_dimensions[
            get_column_letter(
                column
            )
        ].width = width


# ---------------------------------------------------------------------------
# PRODUCT MATCHING
# ---------------------------------------------------------------------------

def _normalized(
    value,
) -> str:
    if value is None:
        return ""

    return str(
        value
    ).strip().casefold()


def _find_matching_row(
    sheet,
    record: PurchaseRecord,
) -> int | None:
    sku = _normalized(
        record.six_digit_sku
    )

    product = _normalized(
        record.product
    )

    store_number = _normalized(
        record.store_number
    )

    for row in range(
        2,
        sheet.max_row + 1,
    ):
        row_sku = _normalized(
            sheet.cell(
                row=row,
                column=SKU_COLUMN,
            ).value
        )

        row_product = _normalized(
            sheet.cell(
                row=row,
                column=PRODUCT_COLUMN,
            ).value
        )

        row_store_number = _normalized(
            sheet.cell(
                row=row,
                column=STORE_NUMBER_COLUMN,
            ).value
        )

        if record.six_digit_sku != NA:
            if (
                row_sku == sku
                and row_store_number
                == store_number
            ):
                return row

        elif (
            row_product == product
            and row_store_number
            == store_number
        ):
            return row

    return None


# ---------------------------------------------------------------------------
# VALUE CONVERSION / CONFLICTS
# ---------------------------------------------------------------------------

def _excel_date(
    value: str,
):
    try:
        return datetime.strptime(
            value,
            "%m/%d/%Y",
        ).date()

    except ValueError:
        return value


def _excel_price(
    value: str,
):
    if value == NA:
        return NA

    try:
        return float(
            value
        )

    except ValueError:
        return value


def _resolve_tax_code_conflict(
    existing: str,
    incoming: str,
    product: str,
) -> str:
    if (
        existing == NA
        and incoming != NA
    ):
        return incoming

    if (
        incoming == NA
        or existing == incoming
    ):
        return existing

    print(
        "\n[WARNING] Tax Code conflict for product:"
        f"\n{product}"
        f"\n1. Keep existing Tax Code: {existing}"
        f"\n2. Use incoming Tax Code: {incoming}"
    )

    while True:
        option = input(
            "\nSelect option: "
        ).strip()

        if option == "1":
            return existing

        if option == "2":
            return incoming

        print(
            "\n[ERROR] Invalid option."
        )


# ---------------------------------------------------------------------------
# PURCHASE APPEND
# ---------------------------------------------------------------------------

def _append_purchase(
    sheet,
    record: PurchaseRecord,
    store_name: str,
) -> bool:
    row = _find_matching_row(
        sheet,
        record,
    )

    created = (
        row is None
    )

    if created:
        row = (
            sheet.max_row + 1
        )

        sheet.cell(
            row=row,
            column=STORE_COLUMN,
            value=(
                record.store
                if record.store != NA
                else store_name
            ),
        )

        sheet.cell(
            row=row,
            column=SKU_COLUMN,
            value=record.six_digit_sku,
        )

        sheet.cell(
            row=row,
            column=PRODUCT_COLUMN,
            value=record.product,
        )

        sheet.cell(
            row=row,
            column=TAX_CODE_COLUMN,
            value=record.tax_code,
        )

        sheet.cell(
            row=row,
            column=STORE_NUMBER_COLUMN,
            value=record.store_number,
        )

        sheet.cell(
            row=row,
            column=COMMON_NAME_COLUMN,
            value=record.common_name,
        )

        sheet.cell(
            row=row,
            column=CATEGORY_COLUMN,
            value=record.category,
        )

    else:
        existing_store = str(
            sheet.cell(
                row=row,
                column=STORE_COLUMN,
            ).value
            or NA
        )

        # Old migrated rows may have Store=NA because the old workbook
        # did not preserve a store name. Once an authoritative receipt type
        # is available, populate the Store column.
        incoming_store = (
            record.store
            if record.store != NA
            else store_name
        )

        if (
            existing_store == NA
            and incoming_store
        ):
            sheet.cell(
                row=row,
                column=STORE_COLUMN,
                value=incoming_store,
            )

        existing_tax = str(
            sheet.cell(
                row=row,
                column=TAX_CODE_COLUMN,
            ).value
            or NA
        )

        selected_tax = (
            _resolve_tax_code_conflict(
                existing=existing_tax,
                incoming=record.tax_code,
                product=record.product,
            )
        )

        sheet.cell(
            row=row,
            column=TAX_CODE_COLUMN,
            value=selected_tax,
        )

        if sheet.cell(
            row=row,
            column=COMMON_NAME_COLUMN,
        ).value in (
            None,
            "",
        ):
            sheet.cell(
                row=row,
                column=COMMON_NAME_COLUMN,
                value=NA,
            )

        if sheet.cell(
            row=row,
            column=CATEGORY_COLUMN,
        ).value in (
            None,
            "",
        ):
            sheet.cell(
                row=row,
                column=CATEGORY_COLUMN,
                value=NA,
            )

    (
        date_column,
        price_column,
    ) = _next_history_pair(
        sheet,
        row,
    )

    date_cell = sheet.cell(
        row=row,
        column=date_column,
    )

    price_cell = sheet.cell(
        row=row,
        column=price_column,
    )

    date_cell.value = _excel_date(
        record.date
    )

    date_cell.number_format = (
        "mm/dd/yyyy"
    )

    price_cell.value = _excel_price(
        record.price
    )

    if record.price != NA:
        price_cell.number_format = (
            "0.00"
        )

    _ensure_total_formula(
        sheet,
        row,
    )

    return created


# ---------------------------------------------------------------------------
# IMPORT AUDIT
# ---------------------------------------------------------------------------

def _source_already_imported(
    workbook,
    source_name: str,
) -> bool:
    sheet = workbook[
        IMPORT_SHEET
    ]

    normalized_source = _normalized(
        source_name
    )

    for row in range(
        2,
        sheet.max_row + 1,
    ):
        existing = sheet.cell(
            row=row,
            column=1,
        ).value

        if (
            _normalized(
                existing
            )
            == normalized_source
        ):
            return True

    return False


def source_already_imported(
    source_path: Path,
) -> bool:
    if not WORKBOOK_PATH.exists():
        return False

    workbook = (
        _load_or_create_workbook()
    )

    return _source_already_imported(
        workbook,
        source_path.name,
    )


def _record_import(
    workbook,
    session: ReceiptSession,
) -> None:
    sheet = workbook[
        IMPORT_SHEET
    ]

    row = (
        sheet.max_row + 1
    )

    sheet.cell(
        row=row,
        column=1,
        value=session.source_path.name,
    )

    sheet.cell(
        row=row,
        column=2,
        value=session.store_number,
    )

    sheet.cell(
        row=row,
        column=3,
        value=_excel_date(
            session.receipt_date
        ),
    )

    sheet.cell(
        row=row,
        column=3,
    ).number_format = (
        "mm/dd/yyyy"
    )

    sheet.cell(
        row=row,
        column=4,
        value=datetime.now(),
    )

    sheet.cell(
        row=row,
        column=4,
    ).number_format = (
        "mm/dd/yyyy hh:mm"
    )


# ---------------------------------------------------------------------------
# SAFE SAVE
# ---------------------------------------------------------------------------

def _atomic_save(
    workbook,
) -> None:
    DATABASE_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    (
        file_descriptor,
        temporary_name,
    ) = tempfile.mkstemp(
        suffix=".xlsx",
        dir=DATABASE_DIR,
    )

    os.close(
        file_descriptor
    )

    temporary_path = Path(
        temporary_name
    )

    try:
        workbook.save(
            temporary_path
        )

        os.replace(
            temporary_path,
            WORKBOOK_PATH,
        )

    finally:
        if temporary_path.exists():
            temporary_path.unlink()


# ---------------------------------------------------------------------------
# PUBLIC COMMIT
# ---------------------------------------------------------------------------

def commit_receipt(
    session: ReceiptSession,
) -> dict:
    workbook = (
        _load_or_create_workbook()
    )

    sheet = workbook[
        PURCHASE_SHEET
    ]

    new_rows = 0
    updated_rows = 0

    for record in (
        session.accepted_purchases
    ):
        created = _append_purchase(
            sheet=sheet,
            record=record,
            store_name=session.receipt_type,
        )

        if created:
            new_rows += 1

        else:
            updated_rows += 1

    # Ensure every populated history row has a Total equation.
    for row in range(
        2,
        sheet.max_row + 1,
    ):
        _ensure_total_formula(
            sheet,
            row,
        )

    _record_import(
        workbook,
        session,
    )

    _format_workbook(
        workbook
    )

    _atomic_save(
        workbook
    )

    return {
        "workbook_path": (
            WORKBOOK_PATH.resolve()
        ),
        "purchases_added": len(
            session.accepted_purchases
        ),
        "existing_histories_updated": (
            updated_rows
        ),
        "new_product_rows": (
            new_rows
        ),
    }
