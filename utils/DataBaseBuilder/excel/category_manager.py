from __future__ import annotations

import os
import tempfile
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from utils.DataBaseBuilder.excel.purchase_history import (
    CATEGORY_COLUMN,
    COMMON_NAME_COLUMN,
    FIXED_HEADERS,
    PURCHASE_SHEET,
    WORKBOOK_PATH,
    _ensure_purchase_schema,
)
from utils.DataBaseBuilder.purchase_record import NA


CATEGORY_MANAGER_SHEET = "Category Manager"
CATEGORY_HEADER = "Category"
SUB_CATEGORY_HEADER = "Sub-Category"
PRODUCT_HEADER_PREFIX = "Product"

HEADER_FILL = "1F4E78"
HEADER_FONT = "FFFFFF"
CATEGORY_FILLS = (
    "D9EAF7",
    "E2F0D9",
    "FFF2CC",
    "FCE4D6",
    "E4DFEC",
    "DDEBF7",
)
BORDER_COLOR = "B7C9D6"


def _text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalized(value) -> str:
    return _text(value).casefold()


def _is_na(value) -> bool:
    return _normalized(value) == _normalized(NA)


def _required_purchase_headers_are_valid(sheet) -> bool:
    actual = [
        _text(sheet.cell(row=1, column=column).value)
        for column in range(1, len(FIXED_HEADERS) + 1)
    ]
    return actual == FIXED_HEADERS


def _load_existing_workbook(
    workbook_path: Path = WORKBOOK_PATH,
):
    workbook_path = Path(workbook_path).expanduser().resolve()

    if not workbook_path.exists():
        raise FileNotFoundError(
            "Purchase History workbook was not found:"
            f"\n{workbook_path}"
        )

    workbook = load_workbook(workbook_path)

    if PURCHASE_SHEET not in workbook.sheetnames:
        workbook.close()
        raise ValueError(
            f'Workbook is missing required sheet: "{PURCHASE_SHEET}".'
        )

    purchase_sheet = workbook[PURCHASE_SHEET]

    # This also performs the safe one-time Category -> Sub-Category
    # header migration without shifting any physical columns.
    _ensure_purchase_schema(purchase_sheet)

    if not _required_purchase_headers_are_valid(purchase_sheet):
        workbook.close()
        raise ValueError(
            "Purchase History has an unrecognized column layout. "
            "Category Manager requires the current ShopGraph Purchase "
            "History schema."
        )

    return workbook


def _read_purchase_history(
    purchase_sheet,
) -> dict:
    """
    Read Common Name -> Sub-Category from Purchase History.

    Common Name is the Category Manager product identity. Multiple Purchase
    History rows may share one Common Name, but they must agree on the detailed
    Sub-Category before Create / Refresh can build a deterministic manager.
    """
    subcategories_by_common_name: dict[str, set[str]] = defaultdict(set)
    rows_by_common_name: dict[str, list[int]] = defaultdict(list)
    invalid_common_name_rows = []

    for row in range(2, purchase_sheet.max_row + 1):
        common_name = _text(
            purchase_sheet.cell(
                row=row,
                column=COMMON_NAME_COLUMN,
            ).value
        )
        subcategory = _text(
            purchase_sheet.cell(
                row=row,
                column=CATEGORY_COLUMN,
            ).value
        )

        # Formula-only / structurally empty trailing rows are ignored.
        product_value = _text(
            purchase_sheet.cell(
                row=row,
                column=5,
            ).value
        )

        if not common_name or _is_na(common_name):
            if product_value:
                invalid_common_name_rows.append(row)
            continue

        if not subcategory:
            subcategory = NA

        subcategories_by_common_name[common_name].add(subcategory)
        rows_by_common_name[common_name].append(row)

    conflicts = {}

    for common_name, subcategories in subcategories_by_common_name.items():
        if len(subcategories) > 1:
            conflicts[common_name] = {
                "subcategories": sorted(
                    subcategories,
                    key=str.casefold,
                ),
                "rows": rows_by_common_name[common_name],
            }

    mapping = {
        common_name: next(iter(subcategories))
        for common_name, subcategories
        in subcategories_by_common_name.items()
        if len(subcategories) == 1
    }

    return {
        "mapping": mapping,
        "rows_by_common_name": dict(rows_by_common_name),
        "invalid_common_name_rows": invalid_common_name_rows,
        "conflicts": conflicts,
    }


def _existing_category_mappings(workbook) -> dict[str, str]:
    """
    Preserve only valid Sub-Category -> Category relationships during refresh.

    Legacy Category Manager sheets used:
        Category | Product 1 | ...

    In that layout the old "Category" is now a Sub-Category, so there is no
    broad Category information to preserve. Only the new two-column hierarchy
    is interpreted as a broad mapping.
    """
    if CATEGORY_MANAGER_SHEET not in workbook.sheetnames:
        return {}

    sheet = workbook[CATEGORY_MANAGER_SHEET]

    if (
        _text(sheet.cell(1, 1).value) != CATEGORY_HEADER
        or _text(sheet.cell(1, 2).value) != SUB_CATEGORY_HEADER
    ):
        return {}

    by_subcategory: dict[str, list[str]] = defaultdict(list)

    for row in range(2, sheet.max_row + 1):
        category = _text(sheet.cell(row, 1).value)
        subcategory = _text(sheet.cell(row, 2).value)

        if (
            not category
            or _is_na(category)
            or not subcategory
            or _is_na(subcategory)
        ):
            continue

        by_subcategory[_normalized(subcategory)].append(category)

    preserved = {}

    for normalized_subcategory, categories in by_subcategory.items():
        distinct = {
            _normalized(category): category
            for category in categories
        }

        if len(distinct) == 1:
            preserved[normalized_subcategory] = next(iter(distinct.values()))

    return preserved


def _subcategory_groups(
    common_to_subcategory: dict[str, str],
) -> list[tuple[str, list[str]]]:
    groups: dict[str, list[str]] = defaultdict(list)

    for common_name, subcategory in common_to_subcategory.items():
        groups[subcategory].append(common_name)

    return [
        (
            subcategory,
            sorted(common_names, key=str.casefold),
        )
        for subcategory, common_names in sorted(
            groups.items(),
            key=lambda item: item[0].casefold(),
        )
    ]


def _delete_manager_if_present(workbook) -> None:
    if CATEGORY_MANAGER_SHEET in workbook.sheetnames:
        workbook.remove(workbook[CATEGORY_MANAGER_SHEET])


def _create_manager_sheet(
    workbook,
    common_to_subcategory: dict[str, str],
    subcategory_to_category: dict[str, str] | None = None,
):
    _delete_manager_if_present(workbook)

    purchase_index = workbook.sheetnames.index(PURCHASE_SHEET)

    sheet = workbook.create_sheet(
        CATEGORY_MANAGER_SHEET,
        purchase_index + 1,
    )

    groups = _subcategory_groups(common_to_subcategory)
    subcategory_to_category = subcategory_to_category or {}

    max_products = max(
        (len(common_names) for _, common_names in groups),
        default=1,
    )

    sheet.cell(row=1, column=1, value=CATEGORY_HEADER)
    sheet.cell(row=1, column=2, value=SUB_CATEGORY_HEADER)

    for product_number in range(1, max_products + 1):
        sheet.cell(
            row=1,
            column=product_number + 2,
            value=f"{PRODUCT_HEADER_PREFIX} {product_number}",
        )

    for row, (subcategory, common_names) in enumerate(groups, start=2):
        category = subcategory_to_category.get(
            subcategory,
            subcategory_to_category.get(
                _normalized(subcategory),
                NA,
            ),
        )

        sheet.cell(row=row, column=1, value=category or NA)
        sheet.cell(row=row, column=2, value=subcategory)

        for column, common_name in enumerate(common_names, start=3):
            sheet.cell(
                row=row,
                column=column,
                value=common_name,
            )

    _format_category_manager(sheet)
    return sheet


def _format_category_manager(sheet) -> None:
    sheet.freeze_panes = "C2"
    sheet.sheet_view.showGridLines = False

    thin = Side(style="thin", color=BORDER_COLOR)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in sheet[1]:
        cell.font = Font(bold=True, color=HEADER_FONT)
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        cell.border = border

    for row in range(2, sheet.max_row + 1):
        fill = PatternFill(
            "solid",
            fgColor=CATEGORY_FILLS[
                (row - 2) % len(CATEGORY_FILLS)
            ],
        )

        for column in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row, column=column)
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

        sheet.cell(row=row, column=1).font = Font(bold=True)
        sheet.cell(row=row, column=2).font = Font(bold=True)

    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 28

    for column in range(3, sheet.max_column + 1):
        sheet.column_dimensions[
            get_column_letter(column)
        ].width = 34

    sheet.row_dimensions[1].height = 24

    for row in range(2, sheet.max_row + 1):
        sheet.row_dimensions[row].height = 36

    sheet.auto_filter.ref = sheet.dimensions


def _manager_structure_errors(sheet) -> list[str]:
    errors = []

    if _text(sheet.cell(1, 1).value) != CATEGORY_HEADER:
        errors.append('Cell A1 must contain "Category".')

    if _text(sheet.cell(1, 2).value) != SUB_CATEGORY_HEADER:
        errors.append('Cell B1 must contain "Sub-Category".')

    if sheet.max_column < 3:
        errors.append(
            "Category Manager must contain at least one Product column."
        )
    else:
        first_product_header = _text(sheet.cell(1, 3).value)
        if not first_product_header.startswith(PRODUCT_HEADER_PREFIX):
            errors.append(
                'Cell C1 must be a Product heading such as "Product 1".'
            )

    return errors


def _parse_manager(sheet) -> dict:
    assignments = []
    blank_category_rows = []
    blank_subcategory_products = []
    subcategory_rows: dict[str, list[dict]] = defaultdict(list)

    for row in range(2, sheet.max_row + 1):
        category = _text(sheet.cell(row, 1).value)
        subcategory = _text(sheet.cell(row, 2).value)

        products = []

        for column in range(3, sheet.max_column + 1):
            common_name = _text(sheet.cell(row, column).value)
            if common_name:
                products.append((common_name, column))

        # Completely empty rows are intentionally ignored.
        if not products and not category and not subcategory:
            continue

        # A row with category/subcategory but no products is an empty taxonomy
        # row. It may be removed during successful cleanup and is ignored here.
        if not products:
            continue

        if not subcategory or _is_na(subcategory):
            for common_name, column in products:
                blank_subcategory_products.append(
                    {
                        "common_name": common_name,
                        "row": row,
                        "column": column,
                    }
                )
            continue

        if not category or _is_na(category):
            blank_category_rows.append(
                {
                    "subcategory": subcategory,
                    "row": row,
                }
            )

        subcategory_rows[_normalized(subcategory)].append(
            {
                "subcategory": subcategory,
                "category": category,
                "row": row,
            }
        )

        for common_name, column in products:
            assignments.append(
                {
                    "common_name": common_name,
                    "subcategory": subcategory,
                    "category": category,
                    "row": row,
                    "column": column,
                }
            )

    return {
        "assignments": assignments,
        "blank_category_rows": blank_category_rows,
        "blank_subcategory_products": blank_subcategory_products,
        "subcategory_rows": dict(subcategory_rows),
    }


def _validate_manager(
    purchase_data: dict,
    manager_data: dict,
    structure_errors: list[str],
) -> dict:
    errors = {
        "structure": list(structure_errors),
        "invalid_purchase_common_names": list(
            purchase_data["invalid_common_name_rows"]
        ),
        "purchase_conflicts": dict(
            purchase_data["conflicts"]
        ),
        "missing": [],
        "unknown": [],
        "duplicates": {},
        "blank_categories": list(
            manager_data["blank_category_rows"]
        ),
        "blank_subcategories": list(
            manager_data["blank_subcategory_products"]
        ),
        "duplicate_subcategories": {},
        "ambiguous_subcategories": {},
    }

    assignments = manager_data["assignments"]
    occurrences: dict[str, list[dict]] = defaultdict(list)

    for assignment in assignments:
        occurrences[assignment["common_name"]].append(assignment)

    expected = set(purchase_data["mapping"])
    found = set(occurrences)

    errors["missing"] = sorted(
        expected - found,
        key=str.casefold,
    )
    errors["unknown"] = sorted(
        found - expected,
        key=str.casefold,
    )
    errors["duplicates"] = {
        common_name: locations
        for common_name, locations in occurrences.items()
        if len(locations) > 1
    }

    for normalized_subcategory, rows in manager_data[
        "subcategory_rows"
    ].items():
        if len(rows) > 1:
            display_name = rows[0]["subcategory"]
            errors["duplicate_subcategories"][display_name] = [
                item["row"]
                for item in rows
            ]

            categories = {
                _normalized(item["category"]): item["category"]
                for item in rows
                if item["category"] and not _is_na(item["category"])
            }

            if len(categories) > 1:
                errors["ambiguous_subcategories"][display_name] = sorted(
                    categories.values(),
                    key=str.casefold,
                )

    has_errors = any(bool(value) for value in errors.values())

    common_to_subcategory = {}
    subcategory_to_category = {}

    if not has_errors:
        common_to_subcategory = {
            assignment["common_name"]: assignment["subcategory"]
            for assignment in assignments
        }
        subcategory_to_category = {
            assignment["subcategory"]: assignment["category"]
            for assignment in assignments
        }

    return {
        "valid": not has_errors,
        "errors": errors,
        "common_to_subcategory": common_to_subcategory,
        "subcategory_to_category": subcategory_to_category,
    }


def _format_validation_report(validation: dict) -> str:
    errors = validation["errors"]
    lines = ["[ERROR] Category Manager validation failed."]

    if errors["structure"]:
        lines.append("\nStructure errors:")
        lines.extend(f"- {message}" for message in errors["structure"])

    if errors["invalid_purchase_common_names"]:
        lines.append(
            "\nPurchase History rows with blank/NA Common Name:"
        )
        lines.extend(
            f"- Row {row}"
            for row in errors["invalid_purchase_common_names"]
        )

    if errors["purchase_conflicts"]:
        lines.append(
            "\nConflicting Purchase History Sub-Categories for one Common Name:"
        )
        for common_name, info in errors["purchase_conflicts"].items():
            lines.append(f'- "{common_name}"')
            lines.append(
                "  Sub-Categories: "
                + ", ".join(info["subcategories"])
            )
            lines.append(
                "  Purchase History rows: "
                + ", ".join(str(row) for row in info["rows"])
            )

    if errors["missing"]:
        lines.append("\nMissing Common Names:")
        lines.extend(
            f"{index}. {common_name}"
            for index, common_name in enumerate(
                errors["missing"],
                start=1,
            )
        )

    if errors["unknown"]:
        lines.append("\nUnknown Common Names:")
        lines.extend(
            f"{index}. {common_name}"
            for index, common_name in enumerate(
                errors["unknown"],
                start=1,
            )
        )

    if errors["duplicates"]:
        lines.append("\nDuplicate Common Name assignments:")
        for common_name, locations in errors["duplicates"].items():
            lines.append(f'\n"{common_name}"')
            for location in locations:
                lines.append(
                    f"    Row {location['row']}: "
                    f"{location['subcategory']}"
                )

    if errors["blank_subcategories"]:
        lines.append("\nProducts assigned to a blank/NA Sub-Category:")
        for location in errors["blank_subcategories"]:
            lines.append(
                f"- {location['common_name']} "
                f"(row {location['row']})"
            )

    if errors["blank_categories"]:
        lines.append("\nSub-Categories without a valid Category:")
        for location in errors["blank_categories"]:
            lines.append(
                f"- {location['subcategory']} "
                f"(row {location['row']})"
            )

    if errors["duplicate_subcategories"]:
        lines.append("\nDuplicate Sub-Category rows:")
        for subcategory, rows in errors[
            "duplicate_subcategories"
        ].items():
            lines.append(
                f'- "{subcategory}" appears on rows: '
                + ", ".join(str(row) for row in rows)
            )

    if errors["ambiguous_subcategories"]:
        lines.append("\nAmbiguous Sub-Category -> Category mappings:")
        for subcategory, categories in errors[
            "ambiguous_subcategories"
        ].items():
            lines.append(
                f'- "{subcategory}" is assigned to: '
                + ", ".join(categories)
            )

    lines.extend(["", "Purchase History was NOT modified."])
    return "\n".join(lines)


def _verified_atomic_save(
    workbook,
    workbook_path: Path = WORKBOOK_PATH,
) -> None:
    workbook_path = Path(workbook_path).expanduser().resolve()
    workbook_path.parent.mkdir(parents=True, exist_ok=True)

    file_descriptor, temporary_name = tempfile.mkstemp(
        suffix=".xlsx",
        dir=workbook_path.parent,
    )
    os.close(file_descriptor)
    temporary_path = Path(temporary_name)

    try:
        workbook.save(temporary_path)

        verification = load_workbook(
            temporary_path,
            read_only=True,
            data_only=False,
        )

        try:
            required = {
                PURCHASE_SHEET,
                CATEGORY_MANAGER_SHEET,
            }
            missing = required - set(verification.sheetnames)

            if missing:
                raise ValueError(
                    "Temporary workbook verification failed. Missing sheet(s): "
                    + ", ".join(sorted(missing))
                )

            verification_purchase = verification[PURCHASE_SHEET]
            if _text(
                verification_purchase.cell(
                    row=1,
                    column=CATEGORY_COLUMN,
                ).value
            ) != "Sub-Category":
                raise ValueError(
                    "Temporary workbook verification failed: "
                    'Purchase History column I is not "Sub-Category".'
                )
        finally:
            verification.close()

        os.replace(temporary_path, workbook_path)

    finally:
        if temporary_path.exists():
            temporary_path.unlink()


def create_or_refresh_category_manager(
    workbook_path: Path = WORKBOOK_PATH,
) -> dict:
    workbook = _load_existing_workbook(workbook_path)

    try:
        purchase_sheet = workbook[PURCHASE_SHEET]
        purchase_data = _read_purchase_history(purchase_sheet)

        if purchase_data["invalid_common_name_rows"]:
            rows = ", ".join(
                str(row)
                for row in purchase_data["invalid_common_name_rows"]
            )
            raise ValueError(
                "Category Manager cannot be created because Purchase History "
                "contains populated rows with blank/NA Common Name. "
                f"Rows: {rows}"
            )

        if purchase_data["conflicts"]:
            details = []

            for common_name, info in purchase_data["conflicts"].items():
                details.append(
                    f"{common_name}: "
                    + ", ".join(info["subcategories"])
                )

            raise ValueError(
                "Category Manager cannot be created because the same Common "
                "Name has conflicting Sub-Categories in Purchase History:\n- "
                + "\n- ".join(details)
            )

        preserved_normalized = _existing_category_mappings(workbook)
        subcategory_to_category = {}

        for subcategory in set(purchase_data["mapping"].values()):
            subcategory_to_category[subcategory] = (
                preserved_normalized.get(
                    _normalized(subcategory),
                    NA,
                )
                if not _is_na(subcategory)
                else NA
            )

        _create_manager_sheet(
            workbook,
            purchase_data["mapping"],
            subcategory_to_category,
        )

        _verified_atomic_save(workbook, workbook_path)

        assigned_category_count = len(
            {
                category
                for category in subcategory_to_category.values()
                if category and not _is_na(category)
            }
        )

        return {
            "workbook_path": Path(workbook_path).resolve(),
            "category_count": assigned_category_count,
            "subcategory_count": len(
                set(purchase_data["mapping"].values())
            ),
            "product_count": len(purchase_data["mapping"]),
        }

    finally:
        workbook.close()


def run_create_or_refresh_category_manager() -> None:
    print("\n=== Create / Refresh Category Manager ===\n")
    print(
        "[INFO] Products and Sub-Categories will be rebuilt from current "
        "Purchase History. Existing valid Sub-Category -> Category mappings "
        "will be preserved; other unapplied layout edits may be discarded."
    )

    try:
        result = create_or_refresh_category_manager()
    except (OSError, ValueError) as error:
        print(f"\n[ERROR] {error}")
        return

    print("\n[OK] Category Manager created / refreshed.")
    print(f"Categories assigned: {result['category_count']}")
    print(f"Sub-Categories: {result['subcategory_count']}")
    print(f"Common Names: {result['product_count']}")
    print(f"Workbook:\n{result['workbook_path']}")


def apply_category_manager(
    workbook_path: Path = WORKBOOK_PATH,
) -> dict:
    workbook = _load_existing_workbook(workbook_path)

    try:
        if CATEGORY_MANAGER_SHEET not in workbook.sheetnames:
            return {
                "success": False,
                "report": (
                    "[ERROR] Category Manager validation failed.\n\n"
                    'Worksheet "Category Manager" was not found.\n\n'
                    "Purchase History was NOT modified."
                ),
            }

        purchase_sheet = workbook[PURCHASE_SHEET]
        manager_sheet = workbook[CATEGORY_MANAGER_SHEET]

        purchase_data = _read_purchase_history(purchase_sheet)
        structure_errors = _manager_structure_errors(manager_sheet)
        manager_data = _parse_manager(manager_sheet)
        validation = _validate_manager(
            purchase_data,
            manager_data,
            structure_errors,
        )

        if not validation["valid"]:
            return {
                "success": False,
                "report": _format_validation_report(validation),
            }

        common_to_subcategory = validation["common_to_subcategory"]
        subcategory_to_category = validation["subcategory_to_category"]
        changed_rows = 0

        # Validation is complete. Only the detailed classification cells in
        # Purchase History may now change. Column I stays physically unchanged.
        for row in range(2, purchase_sheet.max_row + 1):
            common_name = _text(
                purchase_sheet.cell(
                    row=row,
                    column=COMMON_NAME_COLUMN,
                ).value
            )

            if common_name not in common_to_subcategory:
                continue

            subcategory_cell = purchase_sheet.cell(
                row=row,
                column=CATEGORY_COLUMN,
            )
            selected_subcategory = common_to_subcategory[common_name]

            if _text(subcategory_cell.value) != selected_subcategory:
                subcategory_cell.value = selected_subcategory
                changed_rows += 1

        _create_manager_sheet(
            workbook,
            common_to_subcategory,
            subcategory_to_category,
        )

        _verified_atomic_save(workbook, workbook_path)

        return {
            "success": True,
            "workbook_path": Path(workbook_path).resolve(),
            "changed_rows": changed_rows,
            "category_count": len(
                set(subcategory_to_category.values())
            ),
            "subcategory_count": len(subcategory_to_category),
            "product_count": len(common_to_subcategory),
        }

    finally:
        workbook.close()


def run_apply_category_manager() -> None:
    print("\n=== Apply Category Manager to Purchase History ===\n")

    try:
        result = apply_category_manager()
    except (OSError, ValueError) as error:
        print(f"\n[ERROR] {error}")
        return

    if not result["success"]:
        print("\n" + result["report"])
        return

    print("\n[OK] Purchase History Sub-Categories updated successfully.")
    print(f"Purchase History rows changed: {result['changed_rows']}")
    print(f"Categories: {result['category_count']}")
    print(f"Sub-Categories: {result['subcategory_count']}")
    print(f"Common Names: {result['product_count']}")
    print("Category Manager cleaned and synchronized.")
    print(
        "\nRun:\nGenerate / Refresh Purchase Analytics\n"
        "to rebuild Sub Analytics and Analytics using the updated taxonomy."
    )
    print(f"\nWorkbook:\n{result['workbook_path']}")


def load_category_mapping_for_analytics(
    workbook,
) -> dict[str, str]:
    """
    Return a validated Sub-Category -> Category mapping for broad Analytics.

    This intentionally validates the taxonomy columns rather than product
    placement. Broad Analytics reads purchase observations from Purchase
    History and uses Category Manager only for the parent Category relationship.
    """
    if CATEGORY_MANAGER_SHEET not in workbook.sheetnames:
        raise ValueError(
            'Worksheet "Category Manager" was not found. '
            "Create / Refresh Category Manager and assign every Sub-Category "
            "to a Category before generating broad Analytics."
        )

    purchase_sheet = workbook[PURCHASE_SHEET]
    _ensure_purchase_schema(purchase_sheet)

    manager = workbook[CATEGORY_MANAGER_SHEET]

    if (
        _text(manager.cell(1, 1).value) != CATEGORY_HEADER
        or _text(manager.cell(1, 2).value) != SUB_CATEGORY_HEADER
    ):
        raise ValueError(
            'Category Manager must begin with "Category" and "Sub-Category".'
        )

    mapping_by_normalized: dict[str, list[tuple[str, str, int]]] = (
        defaultdict(list)
    )

    for row in range(2, manager.max_row + 1):
        category = _text(manager.cell(row, 1).value)
        subcategory = _text(manager.cell(row, 2).value)

        if not subcategory:
            continue

        mapping_by_normalized[_normalized(subcategory)].append(
            (subcategory, category, row)
        )

    expected_subcategories = set()

    for row in range(2, purchase_sheet.max_row + 1):
        common_name = _text(
            purchase_sheet.cell(row, COMMON_NAME_COLUMN).value
        )
        if not common_name or _is_na(common_name):
            continue

        subcategory = _text(
            purchase_sheet.cell(row, CATEGORY_COLUMN).value
        )

        if not subcategory or _is_na(subcategory):
            raise ValueError(
                "Broad Analytics cannot be generated because Purchase "
                f"History row {row} has no valid Sub-Category."
            )

        expected_subcategories.add(subcategory)

    result = {}
    problems = []

    for subcategory in sorted(expected_subcategories, key=str.casefold):
        matches = mapping_by_normalized.get(
            _normalized(subcategory),
            [],
        )

        if not matches:
            problems.append(
                f'- "{subcategory}" has no Category Manager row.'
            )
            continue

        if len(matches) > 1:
            rows = ", ".join(str(item[2]) for item in matches)
            problems.append(
                f'- "{subcategory}" appears on multiple Category Manager '
                f"rows: {rows}."
            )
            continue

        _, category, row = matches[0]

        if not category or _is_na(category):
            problems.append(
                f'- "{subcategory}" has no valid Category '
                f"(Category Manager row {row})."
            )
            continue

        result[subcategory] = category

    if problems:
        raise ValueError(
            "Broad Analytics requires every current Sub-Category to map "
            "to exactly one Category.\n"
            + "\n".join(problems)
        )

    return result


def display_category_manager_menu() -> None:
    print("\n=== ShopGraph Category Manager ===\n")
    print("1. Create / Refresh Category Manager")
    print("2. Apply Category Manager to Purchase History")
    print("0. Return to Data Base Builder")


def run_category_manager_menu() -> None:
    while True:
        display_category_manager_menu()
        option = input("\nSelect option: ").strip()

        if option == "1":
            run_create_or_refresh_category_manager()
        elif option == "2":
            run_apply_category_manager()
        elif option == "0":
            return
        else:
            print("\n[ERROR] Invalid option.")
