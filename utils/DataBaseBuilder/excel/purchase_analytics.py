from __future__ import annotations

import os
import re
import tempfile
from collections import Counter, defaultdict
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.chart import BarChart, DoughnutChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.marker import DataPoint
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import from_excel

from utils.DataBaseBuilder.excel.purchase_history import (
    PURCHASE_SHEET,
    WORKBOOK_PATH,
    _ensure_purchase_schema,
)
from utils.DataBaseBuilder.excel.category_manager import (
    load_category_mapping_for_analytics,
)


ANALYTICS_SHEET = "Analytics"
ANALYTICS_DATA_SHEET = "_AnalyticsData"
SUB_ANALYTICS_SHEET = "Sub Analytics"
SUB_ANALYTICS_DATA_SHEET = "_SubAnalyticsData"
NA = "NA"

DATE_HEADER_PATTERN = re.compile(r"^Date\s+(\d+)$", re.IGNORECASE)
PRICE_HEADER_PATTERN = re.compile(r"^Price\s+(\d+)$", re.IGNORECASE)

CURRENCY_FORMAT = '$#,##0.00'
DATE_FORMAT = 'mm/dd/yyyy'
MONTH_FORMAT = 'mmm yyyy'

# Approximately twice the original chart dimensions.
CHART_WIDTH = 25.5
CHART_HEIGHT = 14.5
LARGE_CHART_HEIGHT = 16.5

# Accessible, high-contrast palette. Colors intentionally repeat after the
# palette is exhausted; the numeric code remains the primary identifier.
ACCESSIBLE_COLORS = (
    "4472C4",  # blue
    "ED7D31",  # orange
    "70AD47",  # green
    "A5A5A5",  # gray
    "FFC000",  # gold
    "5B9BD5",  # light blue
    "C55A11",  # dark orange
    "548235",  # dark green
    "8064A2",  # purple
    "264478",  # navy
    "9E480E",  # brown
    "636363",  # dark gray
)

# Dashboard regions:
# A:O   -> chart
# Q:W   -> numbered/color key
# Y:AG  -> explanation
CHART_START_COLUMN = 1
KEY_COLOR_COLUMN = 17       # Q
KEY_NUMBER_COLUMN = 18      # R
KEY_DESCRIPTION_COLUMN = 19 # S
KEY_VALUE_COLUMN = 23       # W
OVERVIEW_START_COLUMN = 25  # Y
OVERVIEW_END_COLUMN = 33    # AG


class PurchaseAnalyticsError(ValueError):
    """Raised when Purchase Analytics cannot be generated safely."""


def _normalized_text(value) -> str:
    if value is None:
        return ""

    text = str(value).strip()

    if not text or text.upper() == NA:
        return ""

    return text


def _parse_date(value) -> date | None:
    if value is None:
        return None

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if value <= 0:
            return None

        try:
            converted = from_excel(value)
        except (TypeError, ValueError, OverflowError):
            return None

        if isinstance(converted, datetime):
            return converted.date()

        if isinstance(converted, date):
            return converted

        return None

    text = str(value).strip()

    if not text or text.upper() == NA or text.startswith("="):
        return None

    formats = (
        "%m/%d/%Y",
        "%m/%d/%y",
        "%m-%d-%Y",
        "%m-%d-%y",
        "%Y-%m-%d",
    )

    for format_string in formats:
        try:
            return datetime.strptime(
                text,
                format_string,
            ).date()
        except ValueError:
            continue

    return None


def _parse_price(value) -> float | None:
    if value is None or isinstance(value, bool):
        return None

    if isinstance(value, (int, float, Decimal)):
        try:
            amount = float(value)
        except (TypeError, ValueError, OverflowError):
            return None

        return amount if amount >= 0 else None

    text = str(value).strip()

    if not text or text.upper() == NA or text.startswith("="):
        return None

    text = text.replace("$", "").replace(",", "").strip()

    try:
        amount = float(text)
    except ValueError:
        return None

    return amount if amount >= 0 else None


def _header_map(sheet) -> dict[str, int]:
    headers = {}

    for cell in sheet[1]:
        value = _normalized_text(cell.value)

        if value:
            headers[value] = cell.column

    return headers


def _history_pairs(sheet) -> list[tuple[int, int, int]]:
    dates = {}
    prices = {}

    for cell in sheet[1]:
        header = _normalized_text(cell.value)

        date_match = DATE_HEADER_PATTERN.fullmatch(header)
        if date_match:
            dates[int(date_match.group(1))] = cell.column
            continue

        price_match = PRICE_HEADER_PATTERN.fullmatch(header)
        if price_match:
            prices[int(price_match.group(1))] = cell.column

    pair_numbers = sorted(set(dates) & set(prices))

    return [
        (number, dates[number], prices[number])
        for number in pair_numbers
    ]


def _flatten_purchase_history(sheet) -> tuple[list[dict], dict]:
    headers = _header_map(sheet)
    pairs = _history_pairs(sheet)

    if not pairs:
        return [], {
            "skipped_pairs": 0,
            "history_pairs_found": 0,
        }

    fixed_columns = {
        "store": headers.get("Store"),
        "six_digit_sku": headers.get("Six-Digit SKU"),
        "product": headers.get("Product"),
        "store_number": headers.get("Store Number"),
        "common_name": headers.get("Common Name"),
        "category": headers.get("Sub-Category"),
    }

    observations = []
    skipped_pairs = 0

    for row in range(2, sheet.max_row + 1):
        def value_for(name: str):
            column = fixed_columns.get(name)
            return (
                sheet.cell(row=row, column=column).value
                if column
                else None
            )

        store = _normalized_text(value_for("store")) or "Unknown"
        store_number = _normalized_text(value_for("store_number")) or NA
        product = _normalized_text(value_for("product")) or "Unknown Product"
        common_name = _normalized_text(value_for("common_name")) or product
        category = _normalized_text(value_for("category")) or "Uncategorized"
        sku = _normalized_text(value_for("six_digit_sku")) or NA

        for pair_number, date_column, price_column in pairs:
            raw_date = sheet.cell(
                row=row,
                column=date_column,
            ).value
            raw_price = sheet.cell(
                row=row,
                column=price_column,
            ).value

            has_any_value = (
                _normalized_text(raw_date)
                or _normalized_text(raw_price)
            )

            if not has_any_value:
                continue

            purchase_date = _parse_date(raw_date)
            price = _parse_price(raw_price)

            if purchase_date is None or price is None:
                skipped_pairs += 1
                continue

            observations.append(
                {
                    "date": purchase_date,
                    "price": price,
                    "store": store,
                    "store_number": store_number,
                    "product": product,
                    "common_name": common_name,
                    "category": category,
                    "six_digit_sku": sku,
                    "history_number": pair_number,
                    "source_row": row,
                }
            )

    return observations, {
        "skipped_pairs": skipped_pairs,
        "history_pairs_found": len(pairs),
    }


def _month_start(value: date) -> date:
    return date(value.year, value.month, 1)


def _next_month(value: date) -> date:
    if value.month == 12:
        return date(value.year + 1, 1, 1)

    return date(value.year, value.month + 1, 1)


def _month_range(start: date, end: date) -> list[date]:
    months = []
    current = _month_start(start)
    final = _month_start(end)

    while current <= final:
        months.append(current)
        current = _next_month(current)

    return months


def _sorted_desc(values: dict[str, float | int]) -> list[tuple[str, float | int]]:
    return sorted(
        values.items(),
        key=lambda item: (-item[1], item[0].lower()),
    )


def _numbered_rows(
    rows: list[tuple],
) -> list[tuple]:
    return [
        (index, *row)
        for index, row in enumerate(rows, start=1)
    ]


def _color_for_number(number: int) -> str:
    return ACCESSIBLE_COLORS[
        (number - 1) % len(ACCESSIBLE_COLORS)
    ]


def _aggregate(observations: list[dict]) -> dict:
    total_spending = sum(item["price"] for item in observations)
    earliest = min(item["date"] for item in observations)
    latest = max(item["date"] for item in observations)

    monthly_spending = defaultdict(float)
    category_spending = defaultdict(float)
    store_spending = defaultdict(float)
    product_spending = defaultdict(float)
    category_frequency = Counter()
    monthly_store = defaultdict(lambda: defaultdict(float))

    for item in observations:
        month = _month_start(item["date"])
        price = item["price"]
        category = item["category"]
        store = item["store"]
        product_identity = (
            item["common_name"]
            or item["product"]
            or "Unknown Product"
        )

        monthly_spending[month] += price
        category_spending[category] += price
        store_spending[store] += price
        product_spending[product_identity] += price
        category_frequency[category] += 1
        monthly_store[month][store] += price

    months = _month_range(earliest, latest)

    monthly_rows = [
        (month, monthly_spending.get(month, 0.0))
        for month in months
    ]

    # IMPORTANT: no "Other" bucket. Every category/store is retained so the
    # numbered key can identify every segment individually.
    category_rows = _sorted_desc(category_spending)
    store_rows = _sorted_desc(store_spending)
    share_rows = list(category_rows)
    frequency_rows = _sorted_desc(dict(category_frequency))

    product_rows = _sorted_desc(product_spending)[:10]

    store_names = [
        name
        for name, _ in store_rows
    ]

    monthly_store_rows = []

    for month in months:
        values = [
            monthly_store.get(month, {}).get(
                store,
                0.0,
            )
            for store in store_names
        ]
        monthly_store_rows.append(
            (month, values)
        )

    known_stores = {
        item["store"]
        for item in observations
        if item["store"] != "Unknown"
    }

    known_categories = {
        item["category"]
        for item in observations
        if item["category"] != "Uncategorized"
    }

    if not known_categories:
        known_categories = {
            item["category"]
            for item in observations
        }

    return {
        "total_spending": total_spending,
        "observation_count": len(observations),
        "earliest": earliest,
        "latest": latest,
        "store_count": len(known_stores),
        "category_count": len(known_categories),
        "monthly_spending": monthly_rows,
        "category_spending": category_rows,
        "store_spending": store_rows,
        "category_share": share_rows,
        "product_spending": product_rows,
        "category_frequency": frequency_rows,
        "monthly_store_names": store_names,
        "monthly_store": monthly_store_rows,
    }


def _delete_sheet_if_present(workbook, name: str) -> None:
    if name in workbook.sheetnames:
        workbook.remove(workbook[name])


def _create_output_sheets(
    workbook,
    sheet_name: str,
    data_sheet_name: str,
):
    _delete_sheet_if_present(workbook, sheet_name)
    _delete_sheet_if_present(workbook, data_sheet_name)

    purchase_index = workbook.sheetnames.index(PURCHASE_SHEET)

    analytics = workbook.create_sheet(
        sheet_name,
        purchase_index + 1,
    )
    data_sheet = workbook.create_sheet(
        data_sheet_name,
        purchase_index + 2,
    )

    data_sheet.sheet_state = "hidden"
    analytics.sheet_view.showGridLines = False
    analytics.freeze_panes = "A7"

    return analytics, data_sheet


def _style_dashboard_header(
    analytics,
    aggregates: dict,
) -> None:
    analytics.merge_cells("A1:AG2")

    title = analytics["A1"]
    title.value = "ShopGraph Spending Analytics"
    title.font = Font(
        bold=True,
        size=22,
        color="FFFFFF",
    )
    title.alignment = Alignment(
        horizontal="center",
        vertical="center",
    )
    title.fill = PatternFill(
        "solid",
        fgColor="1F4E78",
    )

    summary = (
        ("Total Spending", aggregates["total_spending"]),
        ("Purchase Observations", aggregates["observation_count"]),
        (
            "Date Range",
            (
                f"{aggregates['earliest'].strftime('%m/%d/%Y')} - "
                f"{aggregates['latest'].strftime('%m/%d/%Y')}"
            ),
        ),
        ("Stores", aggregates["store_count"]),
        ("Categories", aggregates["category_count"]),
    )

    start_columns = (1, 7, 13, 20, 26)

    for (label, value), column in zip(
        summary,
        start_columns,
    ):
        label_cell = analytics.cell(
            row=4,
            column=column,
        )
        value_cell = analytics.cell(
            row=5,
            column=column,
        )

        label_cell.value = label
        label_cell.font = Font(
            bold=True,
            color="404040",
        )
        label_cell.alignment = Alignment(
            horizontal="left"
        )

        value_cell.value = value
        value_cell.font = Font(
            bold=True,
            size=12,
        )
        value_cell.alignment = Alignment(
            horizontal="left"
        )

        if label == "Total Spending":
            value_cell.number_format = CURRENCY_FORMAT

    # Left chart area.
    for column in range(1, 16):
        analytics.column_dimensions[
            get_column_letter(column)
        ].width = 10

    # Spacer.
    analytics.column_dimensions["P"].width = 2

    # Numbered/color key area.
    analytics.column_dimensions["Q"].width = 4
    analytics.column_dimensions["R"].width = 5

    for column in range(19, 23):
        analytics.column_dimensions[
            get_column_letter(column)
        ].width = 15

    analytics.column_dimensions["W"].width = 14
    analytics.column_dimensions["X"].width = 2

    # Explanation area.
    for column in range(25, 34):
        analytics.column_dimensions[
            get_column_letter(column)
        ].width = 13

    analytics.row_dimensions[1].height = 28
    analytics.row_dimensions[2].height = 28


def _write_table(
    sheet,
    *,
    start_column: int,
    start_row: int,
    title: str,
    headers: list[str],
    rows: list[tuple | list],
    currency_columns: set[int] | None = None,
    date_columns: set[int] | None = None,
) -> dict:
    currency_columns = currency_columns or set()
    date_columns = date_columns or set()

    sheet.cell(
        row=start_row,
        column=start_column,
        value=title,
    ).font = Font(bold=True)

    header_row = start_row + 1

    for offset, header in enumerate(headers):
        cell = sheet.cell(
            row=header_row,
            column=start_column + offset,
            value=header,
        )
        cell.font = Font(bold=True)

    data_start = header_row + 1

    for row_offset, values in enumerate(rows):
        row_number = data_start + row_offset

        for column_offset, value in enumerate(values):
            cell = sheet.cell(
                row=row_number,
                column=start_column + column_offset,
                value=value,
            )

            if column_offset in currency_columns:
                cell.number_format = CURRENCY_FORMAT

            if column_offset in date_columns:
                cell.number_format = MONTH_FORMAT

    return {
        "start_column": start_column,
        "header_row": header_row,
        "data_start": data_start,
        "data_end": data_start + len(rows) - 1,
        "column_count": len(headers),
        "row_count": len(rows),
    }


def _apply_point_colors(
    chart,
    count: int,
) -> None:
    if not chart.series:
        return

    points = []

    for index in range(count):
        color = _color_for_number(
            index + 1
        )

        point = DataPoint(idx=index)
        point.graphicalProperties = GraphicalProperties(
            solidFill=color,
        )
        point.graphicalProperties.line.solidFill = color
        points.append(point)

    chart.series[0].dPt = points


def _line_chart(
    data_sheet,
    table: dict,
) -> LineChart:
    chart = LineChart()
    chart.title = "Monthly Spending"
    chart.y_axis.title = "Spending ($)"
    chart.x_axis.title = "Month Number"
    chart.style = 13
    chart.height = CHART_HEIGHT
    chart.width = CHART_WIDTH
    chart.legend = None

    data = Reference(
        data_sheet,
        min_col=table["start_column"] + 2,
        min_row=table["header_row"],
        max_row=table["data_end"],
    )
    categories = Reference(
        data_sheet,
        min_col=table["start_column"],
        min_row=table["data_start"],
        max_row=table["data_end"],
    )

    chart.add_data(
        data,
        titles_from_data=True,
    )
    chart.set_categories(categories)

    if chart.series:
        chart.series[0].graphicalProperties.line.solidFill = (
            ACCESSIBLE_COLORS[0]
        )
        chart.series[0].graphicalProperties.line.width = 28575
        chart.series[0].marker.symbol = "circle"
        chart.series[0].marker.size = 7

    return chart


def _horizontal_bar_chart(
    data_sheet,
    table: dict,
    *,
    title: str,
    x_axis_title: str,
) -> BarChart:
    chart = BarChart()
    chart.type = "bar"
    chart.style = 10
    chart.title = title
    chart.x_axis.title = x_axis_title
    chart.y_axis.title = "Number"
    chart.height = CHART_HEIGHT
    chart.width = CHART_WIDTH
    chart.legend = None
    chart.gapWidth = 45

    data = Reference(
        data_sheet,
        min_col=table["start_column"] + 2,
        min_row=table["header_row"],
        max_row=table["data_end"],
    )
    categories = Reference(
        data_sheet,
        min_col=table["start_column"],
        min_row=table["data_start"],
        max_row=table["data_end"],
    )

    chart.add_data(
        data,
        titles_from_data=True,
    )
    chart.set_categories(categories)

    labels = DataLabelList()
    labels.showVal = True
    chart.dataLabels = labels

    _apply_point_colors(
        chart,
        table["row_count"],
    )

    return chart


def _doughnut_chart(
    data_sheet,
    table: dict,
    title: str = "Spending Share by Category",
) -> DoughnutChart:
    chart = DoughnutChart()
    chart.title = title
    chart.holeSize = 58
    chart.style = 10
    chart.height = LARGE_CHART_HEIGHT
    chart.width = CHART_WIDTH
    chart.legend = None

    data = Reference(
        data_sheet,
        min_col=table["start_column"] + 2,
        min_row=table["header_row"],
        max_row=table["data_end"],
    )
    labels = Reference(
        data_sheet,
        min_col=table["start_column"],
        min_row=table["data_start"],
        max_row=table["data_end"],
    )

    chart.add_data(
        data,
        titles_from_data=True,
    )
    chart.set_categories(labels)

    # The only labels shown on the doughnut are numeric codes.
    label_options = DataLabelList()
    label_options.showCatName = True
    label_options.showPercent = False
    label_options.showVal = False
    label_options.showLeaderLines = True
    chart.dataLabels = label_options

    _apply_point_colors(
        chart,
        table["row_count"],
    )

    return chart


def _stacked_store_chart(
    data_sheet,
    table: dict,
) -> BarChart:
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "stacked"
    chart.overlap = 100
    chart.style = 12
    chart.title = "Monthly Spending by Store"
    chart.y_axis.title = "Spending ($)"
    chart.x_axis.title = "Month Number"
    chart.height = LARGE_CHART_HEIGHT
    chart.width = CHART_WIDTH

    data = Reference(
        data_sheet,
        min_col=table["start_column"] + 1,
        max_col=(
            table["start_column"]
            + table["column_count"]
            - 1
        ),
        min_row=table["header_row"],
        max_row=table["data_end"],
    )
    categories = Reference(
        data_sheet,
        min_col=table["start_column"],
        min_row=table["data_start"],
        max_row=table["data_end"],
    )

    chart.add_data(
        data,
        titles_from_data=True,
    )
    chart.set_categories(categories)

    # Series headers are numeric store codes. Keep the built-in legend because
    # it displays those compact numbers without descriptions.
    for index, series in enumerate(
        chart.series,
        start=1,
    ):
        color = _color_for_number(index)
        series.graphicalProperties.solidFill = color
        series.graphicalProperties.line.solidFill = color

    return chart


def _write_section_header(
    sheet,
    *,
    start_row: int,
    start_column: int,
    end_column: int,
    text: str,
) -> None:
    sheet.merge_cells(
        start_row=start_row,
        start_column=start_column,
        end_row=start_row,
        end_column=end_column,
    )

    cell = sheet.cell(
        row=start_row,
        column=start_column,
    )
    cell.value = text
    cell.font = Font(
        bold=True,
        size=12,
        color="FFFFFF",
    )
    cell.fill = PatternFill(
        "solid",
        fgColor="1F4E78",
    )
    cell.alignment = Alignment(
        horizontal="left",
        vertical="center",
    )


def _format_metric(
    value,
    *,
    kind: str,
) -> str:
    if kind == "currency":
        return f"${float(value):,.2f}"

    if kind == "percent":
        return f"{float(value):.1%}"

    if kind == "count":
        return f"{int(value):,}"

    return str(value)


def _write_numbered_key(
    analytics,
    *,
    start_row: int,
    entries: list[dict],
    title: str,
) -> int:
    _write_section_header(
        analytics,
        start_row=start_row,
        start_column=KEY_COLOR_COLUMN,
        end_column=KEY_VALUE_COLUMN,
        text=title,
    )

    header_row = start_row + 1

    headers = (
        (KEY_COLOR_COLUMN, ""),
        (KEY_NUMBER_COLUMN, "#"),
        (KEY_DESCRIPTION_COLUMN, "Description"),
        (KEY_VALUE_COLUMN, "Value"),
    )

    for column, label in headers:
        cell = analytics.cell(
            row=header_row,
            column=column,
        )
        cell.value = label
        cell.font = Font(bold=True)
        cell.alignment = Alignment(
            horizontal="left",
            vertical="center",
        )

    for offset, entry in enumerate(
        entries,
        start=1,
    ):
        row = header_row + offset
        number = int(entry["number"])
        color = entry.get(
            "color",
            _color_for_number(number),
        )

        color_cell = analytics.cell(
            row=row,
            column=KEY_COLOR_COLUMN,
        )
        color_cell.fill = PatternFill(
            "solid",
            fgColor=color,
        )

        number_cell = analytics.cell(
            row=row,
            column=KEY_NUMBER_COLUMN,
        )
        number_cell.value = number
        number_cell.font = Font(bold=True)
        number_cell.alignment = Alignment(
            horizontal="center",
        )

        analytics.merge_cells(
            start_row=row,
            start_column=KEY_DESCRIPTION_COLUMN,
            end_row=row,
            end_column=KEY_VALUE_COLUMN - 1,
        )

        description_cell = analytics.cell(
            row=row,
            column=KEY_DESCRIPTION_COLUMN,
        )
        description_cell.value = entry["description"]
        description_cell.alignment = Alignment(
            wrap_text=True,
            vertical="top",
        )

        value_cell = analytics.cell(
            row=row,
            column=KEY_VALUE_COLUMN,
        )
        value_cell.value = entry["value"]
        value_cell.alignment = Alignment(
            horizontal="right",
            vertical="top",
        )

        analytics.row_dimensions[row].height = max(
            analytics.row_dimensions[row].height or 15,
            22,
        )

    return header_row + len(entries)


def _write_overview(
    analytics,
    *,
    start_row: int,
    title: str,
    what_it_shows: str,
    how_to_read: str,
    why_useful: str,
    notes: str = "",
) -> int:
    _write_section_header(
        analytics,
        start_row=start_row,
        start_column=OVERVIEW_START_COLUMN,
        end_column=OVERVIEW_END_COLUMN,
        text="How to Read This Graph",
    )

    row = start_row + 2

    sections = (
        ("What it represents", what_it_shows),
        ("How to read it", how_to_read),
        ("What it is useful for", why_useful),
        ("Notes", notes),
    )

    for heading, body in sections:
        if not body:
            continue

        analytics.merge_cells(
            start_row=row,
            start_column=OVERVIEW_START_COLUMN,
            end_row=row,
            end_column=OVERVIEW_END_COLUMN,
        )
        heading_cell = analytics.cell(
            row=row,
            column=OVERVIEW_START_COLUMN,
        )
        heading_cell.value = heading
        heading_cell.font = Font(
            bold=True,
            color="1F4E78",
        )
        row += 1

        analytics.merge_cells(
            start_row=row,
            start_column=OVERVIEW_START_COLUMN,
            end_row=row + 3,
            end_column=OVERVIEW_END_COLUMN,
        )
        body_cell = analytics.cell(
            row=row,
            column=OVERVIEW_START_COLUMN,
        )
        body_cell.value = body
        body_cell.alignment = Alignment(
            wrap_text=True,
            vertical="top",
        )
        row += 5

    return row


def _key_entries_for_months(
    rows: list[tuple[date, float]],
) -> list[dict]:
    return [
        {
            "number": index,
            "color": ACCESSIBLE_COLORS[0],
            "description": month.strftime("%B %Y"),
            "value": _format_metric(
                spending,
                kind="currency",
            ),
        }
        for index, (month, spending)
        in enumerate(rows, start=1)
    ]


def _key_entries_for_named_values(
    rows: list[tuple[str, float | int]],
    *,
    value_kind: str,
    total: float | None = None,
    include_percent: bool = False,
) -> list[dict]:
    entries = []

    for index, (name, value) in enumerate(
        rows,
        start=1,
    ):
        formatted = _format_metric(
            value,
            kind=value_kind,
        )

        if (
            include_percent
            and total
            and total > 0
        ):
            percent = float(value) / total
            formatted = (
                f"{formatted} "
                f"({_format_metric(percent, kind='percent')})"
            )

        entries.append(
            {
                "number": index,
                "color": _color_for_number(index),
                "description": name,
                "value": formatted,
            }
        )

    return entries


def _key_entries_for_stores(
    store_names: list[str],
    store_totals: list[tuple[str, float]],
) -> list[dict]:
    totals = {
        name: value
        for name, value in store_totals
    }

    return [
        {
            "number": index,
            "color": _color_for_number(index),
            "description": store,
            "value": _format_metric(
                totals.get(store, 0.0),
                kind="currency",
            ),
        }
        for index, store
        in enumerate(store_names, start=1)
    ]


def _section_height(
    entry_count: int,
) -> int:
    # Enough vertical space for the 2x chart, key entries, and explanatory text.
    return max(
        36,
        entry_count + 5,
    )


def _add_insufficient_note(
    analytics,
    cell: str,
    message: str,
) -> None:
    analytics[cell] = message
    analytics[cell].font = Font(
        italic=True,
        color="666666",
    )


def _atomic_save(
    workbook,
    workbook_path: Path,
) -> None:
    workbook_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    descriptor, temporary_name = tempfile.mkstemp(
        suffix=".xlsx",
        dir=workbook_path.parent,
    )
    os.close(descriptor)
    temporary_path = Path(temporary_name)

    try:
        workbook.save(temporary_path)
        os.replace(
            temporary_path,
            workbook_path,
        )
    finally:
        if temporary_path.exists():
            temporary_path.unlink()


def _write_chart_section(
    analytics,
    *,
    start_row: int,
    chart,
    entries: list[dict],
    key_title: str,
    overview: dict,
) -> int:
    analytics.add_chart(
        chart,
        f"A{start_row}",
    )

    _write_numbered_key(
        analytics,
        start_row=start_row,
        entries=entries,
        title=key_title,
    )

    _write_overview(
        analytics,
        start_row=start_row,
        title=overview.get("title", ""),
        what_it_shows=overview["what"],
        how_to_read=overview["read"],
        why_useful=overview["useful"],
        notes=overview.get("notes", ""),
    )

    return start_row + _section_height(
        len(entries)
    )


def _replace_dimension_text(
    sheet,
    dimension_label: str,
) -> None:
    if dimension_label == "Category":
        return

    lower_label = dimension_label.lower()
    plural_label = (
        "Sub-Categories"
        if dimension_label == "Sub-Category"
        else f"{dimension_label}s"
    )
    lower_plural = plural_label.lower()

    for row in sheet.iter_rows():
        for cell in row:
            if not isinstance(cell.value, str):
                continue

            text = cell.value

            # Protect text that was already parameterized with Sub-Category so
            # the word "Category" inside "Sub-Category" is not replaced twice.
            text = text.replace(
                plural_label,
                "__SHOPGRAPH_DIMENSION_PLURAL__",
            )
            text = text.replace(
                lower_plural,
                "__shopgraph_dimension_plural__",
            )
            text = text.replace(
                dimension_label,
                "__SHOPGRAPH_DIMENSION__",
            )
            text = text.replace(
                lower_label,
                "__shopgraph_dimension__",
            )

            text = re.sub(
                r"\bCategories\b",
                plural_label,
                text,
            )
            text = re.sub(
                r"\bcategories\b",
                lower_plural,
                text,
            )
            text = re.sub(
                r"\bCategory\b",
                dimension_label,
                text,
            )
            text = re.sub(
                r"\bcategory\b",
                lower_label,
                text,
            )

            text = text.replace(
                "__SHOPGRAPH_DIMENSION_PLURAL__",
                plural_label,
            )
            text = text.replace(
                "__shopgraph_dimension_plural__",
                lower_plural,
            )
            text = text.replace(
                "__SHOPGRAPH_DIMENSION__",
                dimension_label,
            )
            text = text.replace(
                "__shopgraph_dimension__",
                lower_label,
            )
            cell.value = text


def _retitle_dimension_charts(
    analytics,
    dimension_label: str,
) -> None:
    if dimension_label == "Category":
        return

    # Seven existing charts are created in a stable order.
    dimension_titles = {
        1: f"Spending by {dimension_label}",
        3: f"Spending Share by {dimension_label}",
        6: f"Purchase Frequency by {dimension_label}",
    }

    for index, chart in enumerate(analytics._charts):
        if index in dimension_titles:
            chart.title = dimension_titles[index]


def _generate_dashboard(
    workbook,
    observations: list[dict],
    diagnostics: dict,
    *,
    sheet_name: str,
    data_sheet_name: str,
    dimension_label: str,
    dashboard_title: str,
) -> dict:
    analytics, data_sheet = (
        _create_output_sheets(
            workbook,
            sheet_name,
            data_sheet_name,
        )
    )

    if not observations:
        analytics.merge_cells("A1:AG2")
        analytics["A1"] = "ShopGraph Spending Analytics"
        analytics["A1"].font = Font(
            bold=True,
            size=22,
            color="FFFFFF",
        )
        analytics["A1"].fill = PatternFill(
            "solid",
            fgColor="1F4E78",
        )
        analytics["A1"].alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

        analytics["A4"] = (
            "No purchase history is available yet."
        )
        analytics["A4"].font = Font(
            italic=True,
            size=12,
            color="666666",
        )

        data_sheet["A1"] = (
            "No valid purchase observations."
        )
        data_sheet.sheet_state = "hidden"

        analytics["A1"] = dashboard_title
        _replace_dimension_text(
            analytics,
            dimension_label,
        )
        _replace_dimension_text(
            data_sheet,
            dimension_label,
        )

        return {
            "charts_created": 0,
            "purchase_observations": 0,
            "skipped_pairs": diagnostics["skipped_pairs"],
            "history_pairs_found": diagnostics["history_pairs_found"],
            "total_spending": 0.0,
        }

    aggregates = _aggregate(
        observations
    )

    _style_dashboard_header(
        analytics,
        aggregates,
    )

    # ---------------------------------------------------------------
    # Helper tables
    #
    # Every chart uses compact numeric codes instead of verbose category/store/
    # product names. The visible numbered/color key on Analytics explains the
    # codes. This prevents chart labels from becoming unreadable.
    # ---------------------------------------------------------------

    monthly_numbered = [
        (
            index,
            month,
            spending,
        )
        for index, (month, spending)
        in enumerate(
            aggregates["monthly_spending"],
            start=1,
        )
    ]

    category_numbered = [
        (
            index,
            name,
            value,
        )
        for index, (name, value)
        in enumerate(
            aggregates["category_spending"],
            start=1,
        )
    ]

    store_numbered = [
        (
            index,
            name,
            value,
        )
        for index, (name, value)
        in enumerate(
            aggregates["store_spending"],
            start=1,
        )
    ]

    share_numbered = [
        (
            index,
            name,
            value,
        )
        for index, (name, value)
        in enumerate(
            aggregates["category_share"],
            start=1,
        )
    ]

    product_numbered = [
        (
            index,
            name,
            value,
        )
        for index, (name, value)
        in enumerate(
            aggregates["product_spending"],
            start=1,
        )
    ]

    frequency_numbered = [
        (
            index,
            name,
            value,
        )
        for index, (name, value)
        in enumerate(
            aggregates["category_frequency"],
            start=1,
        )
    ]

    monthly_table = _write_table(
        data_sheet,
        start_column=1,
        start_row=1,
        title="Monthly Spending",
        headers=[
            "Number",
            "Month",
            "Spending",
        ],
        rows=monthly_numbered,
        currency_columns={2},
        date_columns={1},
    )

    category_table = _write_table(
        data_sheet,
        start_column=5,
        start_row=1,
        title=f"Spending by {dimension_label}",
        headers=[
            "Number",
            "Category",
            "Spending",
        ],
        rows=category_numbered,
        currency_columns={2},
    )

    store_table = _write_table(
        data_sheet,
        start_column=9,
        start_row=1,
        title="Spending by Store",
        headers=[
            "Number",
            "Store",
            "Spending",
        ],
        rows=store_numbered,
        currency_columns={2},
    )

    share_table = _write_table(
        data_sheet,
        start_column=13,
        start_row=1,
        title="Spending Share by Category",
        headers=[
            "Number",
            "Category",
            "Spending",
        ],
        rows=share_numbered,
        currency_columns={2},
    )

    product_table = _write_table(
        data_sheet,
        start_column=17,
        start_row=1,
        title="Top Products by Spending",
        headers=[
            "Number",
            "Product",
            "Spending",
        ],
        rows=product_numbered,
        currency_columns={2},
    )

    frequency_table = _write_table(
        data_sheet,
        start_column=21,
        start_row=1,
        title=f"Purchase Frequency by {dimension_label}",
        headers=[
            "Number",
            "Category",
            "Purchases",
        ],
        rows=frequency_numbered,
    )

    # Monthly-by-store chart:
    # first column = numbered month code
    # subsequent series headers = numbered store codes
    monthly_store_headers = [
        "Month Number",
        *[
            str(index)
            for index, _
            in enumerate(
                aggregates[
                    "monthly_store_names"
                ],
                start=1,
            )
        ],
    ]

    monthly_store_rows = [
        [
            month_index,
            *values,
        ]
        for month_index, (_, values)
        in enumerate(
            aggregates["monthly_store"],
            start=1,
        )
    ]

    monthly_store_table = _write_table(
        data_sheet,
        start_column=25,
        start_row=1,
        title="Monthly Spending by Store",
        headers=monthly_store_headers,
        rows=monthly_store_rows,
        currency_columns=set(
            range(
                1,
                len(monthly_store_headers),
            )
        ),
    )

    charts_created = 0
    current_row = 8

    # ---------------------------------------------------------------
    # 1. MONTHLY SPENDING
    # ---------------------------------------------------------------
    if monthly_table["row_count"] > 0:
        chart = _line_chart(
            data_sheet,
            monthly_table,
        )

        entries = _key_entries_for_months(
            aggregates["monthly_spending"]
        )

        current_row = _write_chart_section(
            analytics,
            start_row=current_row,
            chart=chart,
            entries=entries,
            key_title="Number / Month",
            overview={
                "what": (
                    "Shows total household spending for each calendar month. "
                    "Every valid Date N / Price N purchase observation is added "
                    "to its month."
                ),
                "read": (
                    "The x-axis uses month numbers instead of month names to "
                    "keep the graph clear. Match each number to the month and "
                    "dollar total in the key. A rising line means monthly "
                    "spending increased; a falling line means it decreased."
                ),
                "useful": (
                    "Use this to identify overall spending trends, unusually "
                    "expensive months, and whether spending is gradually moving "
                    "up or down over time."
                ),
                "notes": (
                    "Months with no spending between the earliest and latest "
                    "purchase dates are included as zero-spend months."
                ),
            },
        )
        charts_created += 1

    # ---------------------------------------------------------------
    # 2. SPENDING BY CATEGORY
    # ---------------------------------------------------------------
    if category_table["row_count"] > 0:
        chart = _horizontal_bar_chart(
            data_sheet,
            category_table,
            title=f"Spending by {dimension_label}",
            x_axis_title="Spending ($)",
        )

        entries = _key_entries_for_named_values(
            aggregates["category_spending"],
            value_kind="currency",
        )

        current_row = _write_chart_section(
            analytics,
            start_row=current_row,
            chart=chart,
            entries=entries,
            key_title="Number / Category",
            overview={
                "what": (
                    "Compares total spending across every category in Purchase "
                    "History. No categories are collapsed into an 'Other' group."
                ),
                "read": (
                    "Each horizontal bar is identified by a number and color. "
                    "Use the key to map that number/color to the category name "
                    "and exact spending amount. Longer bars mean more money was "
                    "spent in that category."
                ),
                "useful": (
                    "Use this to see which types of goods are the biggest "
                    "drivers of household spending."
                ),
            },
        )
        charts_created += 1

    # ---------------------------------------------------------------
    # 3. SPENDING BY STORE
    # ---------------------------------------------------------------
    if store_table["row_count"] > 0:
        chart = _horizontal_bar_chart(
            data_sheet,
            store_table,
            title="Spending by Store",
            x_axis_title="Spending ($)",
        )

        entries = _key_entries_for_named_values(
            aggregates["store_spending"],
            value_kind="currency",
        )

        current_row = _write_chart_section(
            analytics,
            start_row=current_row,
            chart=chart,
            entries=entries,
            key_title="Number / Store",
            overview={
                "what": (
                    "Compares how much money was spent at each retailer across "
                    "the entire Purchase History."
                ),
                "read": (
                    "Each store is represented by a numbered colored bar. Match "
                    "the number/color to the key. Longer bars indicate more "
                    "total spending at that store."
                ),
                "useful": (
                    "Use this to understand which retailers receive the largest "
                    "share of your household shopping budget."
                ),
                "notes": (
                    "Store Number is intentionally ignored here; purchases are "
                    "grouped by the Store name."
                ),
            },
        )
        charts_created += 1

    # ---------------------------------------------------------------
    # 4. CATEGORY SHARE
    # ---------------------------------------------------------------
    if share_table["row_count"] > 0:
        chart = _doughnut_chart(
            data_sheet,
            share_table,
            title=f"Spending Share by {dimension_label}",
        )

        entries = _key_entries_for_named_values(
            aggregates["category_share"],
            value_kind="currency",
            total=aggregates["total_spending"],
            include_percent=True,
        )

        current_row = _write_chart_section(
            analytics,
            start_row=current_row,
            chart=chart,
            entries=entries,
            key_title="Number / Category",
            overview={
                "what": (
                    "Shows each category's share of total spending as part of "
                    "the whole shopping budget."
                ),
                "read": (
                    "The doughnut itself displays only numeric category codes. "
                    "Match each number/color to the key for the category name, "
                    "dollar amount, and percentage of total spending."
                ),
                "useful": (
                    "Use this for a fast visual picture of the household "
                    "spending mix—especially which categories take the largest "
                    "portion of the budget."
                ),
                "notes": (
                    "Unlike the older dashboard, this graph does not create a "
                    "large 'Other' slice. Every category receives its own "
                    "numbered segment."
                ),
            },
        )
        charts_created += 1

    # ---------------------------------------------------------------
    # 5. TOP PRODUCTS
    # ---------------------------------------------------------------
    if product_table["row_count"] > 0:
        chart = _horizontal_bar_chart(
            data_sheet,
            product_table,
            title="Top 10 Products by Spending",
            x_axis_title="Spending ($)",
        )

        entries = _key_entries_for_named_values(
            aggregates["product_spending"],
            value_kind="currency",
        )

        current_row = _write_chart_section(
            analytics,
            start_row=current_row,
            chart=chart,
            entries=entries,
            key_title="Number / Product",
            overview={
                "what": (
                    "Ranks the ten individual product identities with the most "
                    "accumulated spending across all recorded purchases."
                ),
                "read": (
                    "Each product is represented by a numbered colored bar. "
                    "Use the key to see the product name and its total spending. "
                    "The longest bar is the largest individual spending driver."
                ),
                "useful": (
                    "Use this to identify specific products that contribute the "
                    "most to total spending and may be worth price-comparing or "
                    "watching over time."
                ),
                "notes": (
                    "Product identity prefers Common Name and falls back to "
                    "Product when Common Name is unavailable."
                ),
            },
        )
        charts_created += 1

    # ---------------------------------------------------------------
    # 6. MONTHLY SPENDING BY STORE
    # ---------------------------------------------------------------
    if (
        monthly_store_table["row_count"] > 0
        and aggregates["monthly_store_names"]
    ):
        chart = _stacked_store_chart(
            data_sheet,
            monthly_store_table,
        )

        store_entries = _key_entries_for_stores(
            aggregates["monthly_store_names"],
            aggregates["store_spending"],
        )

        # Add a second compact month-number mapping below the store mapping.
        month_entries = _key_entries_for_months(
            aggregates["monthly_spending"]
        )

        combined_entries = list(
            store_entries
        )

        # The chart series numbers refer to stores. Months are x-axis numbers.
        # Month rows are appended with the same neutral color and descriptions
        # prefixed so the two number systems are unambiguous.
        for entry in month_entries:
            combined_entries.append(
                {
                    "number": entry["number"],
                    "color": "D9EAF7",
                    "description": (
                        "Month "
                        + entry["description"]
                    ),
                    "value": entry["value"],
                }
            )

        current_row = _write_chart_section(
            analytics,
            start_row=current_row,
            chart=chart,
            entries=combined_entries,
            key_title="Store Numbers + Month Numbers",
            overview={
                "what": (
                    "Shows how monthly spending is divided among stores. Each "
                    "column is one month; each colored section inside the "
                    "column represents one store."
                ),
                "read": (
                    "Store series use the numbered color key. The x-axis uses "
                    "month numbers. Match those month numbers to the month rows "
                    "in the key. The total height of a column is total monthly "
                    "spending; the colored portions show which stores made up "
                    "that total."
                ),
                "useful": (
                    "Use this to see whether changes in monthly spending are "
                    "connected to a particular store and how your retailer mix "
                    "changes over time."
                ),
                "notes": (
                    "All stores are retained. There is no catch-all 'Other' "
                    "store series."
                ),
            },
        )
        charts_created += 1

    # ---------------------------------------------------------------
    # 7. PURCHASE FREQUENCY BY CATEGORY
    # ---------------------------------------------------------------
    if frequency_table["row_count"] > 0:
        chart = _horizontal_bar_chart(
            data_sheet,
            frequency_table,
            title=f"Purchase Frequency by {dimension_label}",
            x_axis_title="Purchases",
        )

        entries = _key_entries_for_named_values(
            aggregates["category_frequency"],
            value_kind="count",
        )

        current_row = _write_chart_section(
            analytics,
            start_row=current_row,
            chart=chart,
            entries=entries,
            key_title="Number / Category",
            overview={
                "what": (
                    "Counts how many valid purchase observations occurred in "
                    "each category. This measures frequency, not dollars."
                ),
                "read": (
                    "Each category is a numbered colored bar. Match the number "
                    "and color to the key. A longer bar means the category was "
                    "purchased more frequently."
                ),
                "useful": (
                    "Use this alongside Spending by Category to distinguish "
                    "frequent routine purchases from categories that are bought "
                    "less often but cost more."
                ),
            },
        )
        charts_created += 1

    # Keep helper data available for debugging but invisible during normal use.
    data_sheet.sheet_state = "hidden"

    analytics["A1"] = dashboard_title
    _replace_dimension_text(
        analytics,
        dimension_label,
    )
    _replace_dimension_text(
        data_sheet,
        dimension_label,
    )
    return {
        "charts_created": charts_created,
        "purchase_observations": aggregates[
            "observation_count"
        ],
        "total_spending": aggregates[
            "total_spending"
        ],
        "skipped_pairs": diagnostics[
            "skipped_pairs"
        ],
        "history_pairs_found": diagnostics[
            "history_pairs_found"
        ],
    }


def _broad_category_observations(
    observations: list[dict],
    subcategory_to_category: dict[str, str],
) -> list[dict]:
    broad = []

    for observation in observations:
        subcategory = observation["category"]

        if subcategory not in subcategory_to_category:
            raise PurchaseAnalyticsError(
                f'Sub-Category "{subcategory}" does not have a broad Category.'
            )

        item = dict(observation)
        item["category"] = subcategory_to_category[
            subcategory
        ]
        broad.append(item)

    return broad


def generate_purchase_analytics(
    workbook_path: Path = WORKBOOK_PATH,
) -> dict:
    """
    Refresh both analytical layers.

    Sub Analytics is the existing detailed dashboard and reads Sub-Category
    directly from Purchase History. Analytics is the broad dashboard and joins
    Purchase History Sub-Category to Category Manager's Category mapping.
    """
    workbook_path = Path(
        workbook_path
    ).expanduser().resolve()

    if not workbook_path.exists():
        raise FileNotFoundError(
            "Purchase History workbook does not exist: "
            f"{workbook_path}"
        )

    workbook = load_workbook(
        workbook_path,
        data_only=False,
    )

    try:
        if PURCHASE_SHEET not in workbook.sheetnames:
            raise PurchaseAnalyticsError(
                f"Workbook does not contain '{PURCHASE_SHEET}'."
            )

        purchase_sheet = workbook[
            PURCHASE_SHEET
        ]
        _ensure_purchase_schema(
            purchase_sheet
        )

        observations, diagnostics = (
            _flatten_purchase_history(
                purchase_sheet
            )
        )

        sub_summary = _generate_dashboard(
            workbook,
            observations,
            diagnostics,
            sheet_name=SUB_ANALYTICS_SHEET,
            data_sheet_name=SUB_ANALYTICS_DATA_SHEET,
            dimension_label="Sub-Category",
            dashboard_title="ShopGraph Sub Analytics",
        )

        broad_summary = None
        broad_error = None

        try:
            subcategory_to_category = (
                load_category_mapping_for_analytics(
                    workbook
                )
            )
            broad_observations = (
                _broad_category_observations(
                    observations,
                    subcategory_to_category,
                )
            )

            broad_summary = _generate_dashboard(
                workbook,
                broad_observations,
                diagnostics,
                sheet_name=ANALYTICS_SHEET,
                data_sheet_name=ANALYTICS_DATA_SHEET,
                dimension_label="Category",
                dashboard_title="ShopGraph Spending Analytics",
            )

        except (
            ValueError,
            PurchaseAnalyticsError,
        ) as error:
            broad_error = str(error)

            # Never leave a stale broad dashboard after the hierarchy becomes
            # incomplete or invalid.
            _delete_sheet_if_present(
                workbook,
                ANALYTICS_SHEET,
            )
            _delete_sheet_if_present(
                workbook,
                ANALYTICS_DATA_SHEET,
            )

        _atomic_save(
            workbook,
            workbook_path,
        )

        return {
            "workbook_path": workbook_path,
            "sub_analytics": {
                "success": True,
                **sub_summary,
            },
            "analytics": (
                {
                    "success": True,
                    **broad_summary,
                }
                if broad_summary is not None
                else {
                    "success": False,
                    "error": (
                        broad_error
                        or "Broad Analytics could not be generated."
                    ),
                }
            ),
        }

    finally:
        workbook.close()
