from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import load_workbook

from utils.constants import DATA_DIR


DATABASE_DIR = DATA_DIR / "database"

WORKBOOK_PATH = (
    DATABASE_DIR
    / "shopgraph_purchase_history.xlsx"
)

OUTPUT_PATH = (
    DATABASE_DIR
    / "shopgraph_category_manager.txt"
)

CATEGORY_MANAGER_SHEET = "Category Manager"


def _cell_to_text(value) -> str:
    if value is None:
        return ""

    return str(value)


def export_category_manager_as_csv_txt() -> Path:
    """
    Export the Category Manager worksheet as CSV-formatted plain text.

    Source:
        data/database/shopgraph_purchase_history.xlsx

    Worksheet:
        Category Manager

    Output:
        data/database/shopgraph_category_manager.txt

    Formula cells are preserved as formula text.
    """
    if not WORKBOOK_PATH.exists():
        raise FileNotFoundError(
            "Purchase History workbook was not found:"
            f"\n{WORKBOOK_PATH.resolve()}"
        )

    workbook = load_workbook(
        WORKBOOK_PATH,
        data_only=False,
        read_only=True,
    )

    try:
        if CATEGORY_MANAGER_SHEET not in workbook.sheetnames:
            raise ValueError(
                f'Worksheet "{CATEGORY_MANAGER_SHEET}" '
                f"was not found in {WORKBOOK_PATH.name}. "
                "Create / Refresh Category Manager first."
            )

        worksheet = workbook[
            CATEGORY_MANAGER_SHEET
        ]

        DATABASE_DIR.mkdir(
            parents=True,
            exist_ok=True,
        )

        with OUTPUT_PATH.open(
            "w",
            encoding="utf-8",
            newline="",
        ) as output_file:
            writer = csv.writer(
                output_file,
                dialect="excel",
                quoting=csv.QUOTE_MINIMAL,
                lineterminator="\n",
            )

            for row in worksheet.iter_rows(
                values_only=True,
            ):
                writer.writerow(
                    [
                        _cell_to_text(value)
                        for value in row
                    ]
                )

    finally:
        workbook.close()

    return OUTPUT_PATH.resolve()


def run_category_manager_txt_export() -> None:
    print(
        "\n=== Export Category Manager as CSV TXT ===\n"
    )

    try:
        output_path = (
            export_category_manager_as_csv_txt()
        )
    except (
        FileNotFoundError,
        ValueError,
        PermissionError,
        OSError,
    ) as error:
        print(
            f"[ERROR] {error}"
        )
        return

    print(
        "[OK] Category Manager exported "
        "as CSV-formatted text."
    )

    print(
        "\nSource:"
        f"\n{WORKBOOK_PATH.resolve()}"
    )

    print(
        "\nWorksheet:"
        f"\n{CATEGORY_MANAGER_SHEET}"
    )

    print(
        "\nOutput:"
        f"\n{output_path}"
    )


if __name__ == "__main__":
    run_category_manager_txt_export()
