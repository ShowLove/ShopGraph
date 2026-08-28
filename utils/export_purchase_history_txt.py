from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import load_workbook

from utils.constants import DATA_DIR


DATABASE_DIR = DATA_DIR / "database"

WORKBOOK_PATH = (
    DATABASE_DIR
    / "shopgraph_purchase_history.xlsx"
)

OUTPUT_PATH = (
    DATABASE_DIR
    / "shopgraph_purchase_history.txt"
)

PURCHASE_HISTORY_SHEET = "Purchase History"


def _cell_to_text(value) -> str:
    if value is None:
        return ""

    return str(value)


def export_purchase_history_as_csv_txt() -> Path:
    """
    Export the Purchase History worksheet as CSV-formatted plain text.

    Source:
        data/database/shopgraph_purchase_history.xlsx

    Output:
        data/database/shopgraph_purchase_history.txt

    Formula cells are preserved as formula text. This avoids silently losing
    formula content when a cached Excel-calculated value is unavailable.
    """
    if not WORKBOOK_PATH.exists():
        raise FileNotFoundError(
            "Purchase History workbook was not found:"
            f"\n{WORKBOOK_PATH.resolve()}"
        )

    workbook = load_workbook(
        WORKBOOK_PATH,
        data_only=False,
        read_only=True,
    )

    try:
        if PURCHASE_HISTORY_SHEET not in workbook.sheetnames:
            raise ValueError(
                f'Worksheet "{PURCHASE_HISTORY_SHEET}" '
                f"was not found in {WORKBOOK_PATH.name}."
            )

        worksheet = workbook[
            PURCHASE_HISTORY_SHEET
        ]

        DATABASE_DIR.mkdir(
            parents=True,
            exist_ok=True,
        )

        with OUTPUT_PATH.open(
            "w",
            encoding="utf-8",
            newline="",
        ) as output_file:
            writer = csv.writer(
                output_file,
                dialect="excel",
                quoting=csv.QUOTE_MINIMAL,
                lineterminator="\n",
            )

            for row in worksheet.iter_rows(
                values_only=True,
            ):
                writer.writerow(
                    [
                        _cell_to_text(value)
                        for value in row
                    ]
                )

    finally:
        workbook.close()

    return OUTPUT_PATH.resolve()


def run_purchase_history_txt_export() -> None:
    print(
        "\n=== Export Purchase History as CSV TXT ===\n"
    )

    try:
        output_path = (
            export_purchase_history_as_csv_txt()
        )
    except (
        FileNotFoundError,
        ValueError,
        PermissionError,
        OSError,
    ) as error:
        print(
            f"[ERROR] {error}"
        )
        return

    print(
        "[OK] Purchase History exported "
        "as CSV-formatted text."
    )

    print(
        "\nSource:"
        f"\n{WORKBOOK_PATH.resolve()}"
    )

    print(
        "\nOutput:"
        f"\n{output_path}"
    )


if __name__ == "__main__":
    run_purchase_history_txt_export()
